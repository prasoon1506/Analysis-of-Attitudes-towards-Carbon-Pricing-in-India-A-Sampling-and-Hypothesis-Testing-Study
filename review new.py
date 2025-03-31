from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import re

def is_percentage(value):
    if isinstance(value, str):
        cleaned = value.strip().rstrip('%')
        try:
            float(cleaned)
            return True
        except ValueError:
            return False
    return isinstance(value, (int, float))

def get_numeric_value(value):
    if isinstance(value, str):
        return float(value.strip().rstrip('%'))
    return float(value)

def format_excel_file(input_file, output_file):
    wb = load_workbook(input_file)
    ws = wb.active
    
    # Using the exact colors from the image
    light_red_fill = PatternFill(start_color='FFFFD5D5', end_color='FFFFD5D5', fill_type='solid')  # Light pink/red from image
    light_green_fill = PatternFill(start_color='FFD5FFDA', end_color='FFD5FFDA', fill_type='solid')  # Light green from image
    header_fill = PatternFill(start_color='FF4F81BD', end_color='FF4F81BD', fill_type='solid')  # Professional blue
    
    # Border styles
    thin_border = Side(border_style="thin", color="FF000000")
    medium_border = Side(border_style="medium", color="FF000000")
    border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    header_border = Border(left=thin_border, right=thin_border, top=medium_border, bottom=medium_border)
    
    # Font styles
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFFFF")  # White text for headers
    bold_font = Font(name='Calibri', size=11, bold=True)
    regular_font = Font(name='Calibri', size=11)
    
    # Format header row
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = header_border
    
    # Adjust column widths for better readability
    for col in range(1, ws.max_column + 1):
        column_letter = ws.cell(row=1, column=col).column_letter
        max_length = 0
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = max(max_length + 2, 12)  # Minimum width of 12
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Format all cells with borders and alignment
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = regular_font
    
    # Apply comparison formatting with the exact colors from the image
    column_pairs = [('G', 'H'), ('I', 'J'), ('K', 'L'), ('M', 'N')]
    
    for row in range(2, ws.max_row + 1):
        for col1, col2 in column_pairs:
            cell1 = ws[f'{col1}{row}']
            cell2 = ws[f'{col2}{row}']
            
            if cell1.value is None or cell2.value is None:
                continue
                
            if is_percentage(cell1.value) and is_percentage(cell2.value):
                val1 = get_numeric_value(cell1.value)
                val2 = get_numeric_value(cell2.value)
                
                # Apply the exact colors from the image
                if val2 < val1:
                    ws[f'{col2}{row}'].fill = light_red_fill
                    # Red text color to match image
                    ws[f'{col2}{row}'].font = Font(name='Calibri', size=11, bold=True, color='FF800000')
                else:
                    ws[f'{col2}{row}'].fill = light_green_fill
                    # Green text color to match image
                    ws[f'{col2}{row}'].font = Font(name='Calibri', size=11, bold=True, color='FF008000')
                
                # Make percentage values bold
                ws[f'{col1}{row}'].font = bold_font
                
                # Ensure percentage format is applied
                if isinstance(cell1.value, (int, float)):
                    ws[f'{col1}{row}'].number_format = '0.00%'
                if isinstance(cell2.value, (int, float)):
                    ws[f'{col2}{row}'].number_format = '0.00%'
    
    # Add alternating row colors for better readability
    for row in range(2, ws.max_row + 1, 2):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            current_fill = cell.fill.start_color.index
            # Only apply alternating color if cell doesn't already have red or green fill
            if current_fill != 'FFFFD5D5' and current_fill != 'FFD5FFDA':
                cell.fill = PatternFill(start_color='FFF5F5F5', end_color='FFF5F5F5', fill_type='solid')
    
    # Freeze the top row
    ws.freeze_panes = 'A2'
    
    wb.save(output_file)
    print(f"File has been processed and saved as {output_file}")

if __name__ == "__main__":
    input_file = "/content/Copy of President_Report_YTD.xlsx"
    output_file = "output ytd.xlsx"
    format_excel_file(input_file, output_file)
