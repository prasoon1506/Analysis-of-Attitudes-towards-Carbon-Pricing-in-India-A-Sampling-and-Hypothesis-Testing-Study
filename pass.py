from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, GradientFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
import re
def is_percentage(value):
    if isinstance(value, str):
        cleaned = value.strip().rstrip('%')
        try:
            float(cleaned)
            return True
        except ValueError:
            return False
    return isinstance(value, (int, float))
def get_numeric_value(value):
    if isinstance(value, str):
        return float(value.strip().rstrip('%'))
    return float(value)
def format_excel_file(input_file, output_file):
    wb = load_workbook(input_file)
    ws = wb.active
    colors = {'excellent': PatternFill(start_color='FF1B8F47', end_color='FF1B8F47', fill_type='solid'),'good': PatternFill(start_color='FF66BB6A', end_color='FF66BB6A', fill_type='solid'),'warning': PatternFill(start_color='FFFFA726', end_color='FFFFA726', fill_type='solid'),'critical': PatternFill(start_color='FFF44336', end_color='FFF44336', fill_type='solid'),'poor': PatternFill(start_color='FFEF5350', end_color='FFEF5350', fill_type='solid'),'header_primary': PatternFill(start_color='FF1565C0', end_color='FF1565C0', fill_type='solid'),'header_secondary': PatternFill(start_color='FF1976D2', end_color='FF1976D2', fill_type='solid'),'stripe_light': PatternFill(start_color='FFF8F9FA', end_color='FFF8F9FA', fill_type='solid'),'stripe_medium': PatternFill(start_color='FFE3F2FD', end_color='FFE3F2FD', fill_type='solid'),}
    borders = {'thin': Side(border_style="thin", color="FFE0E0E0"),'medium': Side(border_style="medium", color="FF1565C0"),'thick': Side(border_style="thick", color="FF0D47A1"),'double': Side(border_style="double", color="FF1565C0"),}
    header_border = Border(left=borders['medium'], right=borders['medium'],top=borders['thick'],bottom=borders['double'])
    data_border = Border(left=borders['thin'],right=borders['thin'],top=borders['thin'],bottom=borders['thin'])
    fonts = {'header': Font(name='Segoe UI', size=12, bold=True, color="FFFFFFFF"),'subheader': Font(name='Segoe UI', size=11, bold=True, color="FF1565C0"),'data_bold': Font(name='Segoe UI', size=10, bold=True, color="FF212121"),'data_regular': Font(name='Segoe UI', size=10, color="FF424242"),'success': Font(name='Segoe UI', size=10, bold=True, color="FF1B5E20"),'warning': Font(name='Segoe UI', size=10, bold=True, color="FFF57F17"),'error': Font(name='Segoe UI', size=10, bold=True, color="FFB71C1C"),}
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = colors['header_primary']
        cell.font = fonts['header']
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True,text_rotation=0,indent=0)
        cell.border = header_border
    for col in range(1, ws.max_column + 1):
        column_letter = ws.cell(row=1, column=col).column_letter
        max_length = 0
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                display_length = len(str(cell_value))
                if is_percentage(cell_value):
                    display_length += 2  # Extra space for % and formatting
                max_length = max(max_length, display_length)
        adjusted_width = max(min(max_length + 3, 25), 10)
        ws.column_dimensions[column_letter].width = adjusted_width
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = data_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = fonts['data_regular']
    column_pairs = [('P', 'Q'), ('R', 'S'), ('T', 'U')]
    for row in range(2, ws.max_row + 1):
        for col1, col2 in column_pairs:
            cell1 = ws[f'{col1}{row}']
            cell2 = ws[f'{col2}{row}']
            if cell1.value is None or cell2.value is None:
                continue
            if is_percentage(cell1.value) and is_percentage(cell2.value):
                val1 = get_numeric_value(cell1.value)
                val2 = get_numeric_value(cell2.value)
                if val1 > 0:
                    diff_percentage = ((val2 - val1) / val1) * 100
                else:
                    diff_percentage = 0
                if val2 >= val1 * 1.1:
                    cell2.fill = colors['excellent']
                    cell2.font = fonts['success']
                elif val2 >= val1:
                    cell2.fill = colors['good']
                    cell2.font = fonts['success']
                elif val2 >= val1 * 0.95:
                    cell2.fill = colors['warning']
                    cell2.font = fonts['warning']
                elif val2 >= val1 * 0.85:
                    cell2.fill = colors['poor']
                    cell2.font = fonts['error']
                else:
                    cell2.fill = colors['critical']
                    cell2.font = fonts['error']
                cell1.font = fonts['data_bold']
                cell1.fill = colors['stripe_medium']
                if isinstance(cell1.value, (int, float)):
                    cell1.number_format = '0.00%'
                if isinstance(cell2.value, (int, float)):
                    cell2.number_format = '0.00%'
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            current_fill = getattr(cell.fill, 'start_color', None)
            if (current_fill is None or 
                current_fill.index in ['00000000', 'FFFFFFFF'] or 
                getattr(current_fill, 'index', None) in ['00000000', 'FFFFFFFF']):
                if row % 2 == 0:
                    cell.fill = colors['stripe_light']
                else:
                    pass
    ws.freeze_panes = 'A2'
    if ws.max_row > 1 and ws.max_column > 1:
        ws.print_area = f'A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}'
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    wb.save(output_file)
    print(f"✨ Enhanced file has been processed and saved as {output_file}")
    print("🎨 Applied modern styling with:")
    print("   • Professional color scheme")
    print("   • Enhanced performance indicators")
    print("   • Modern typography")
    print("   • Improved visual hierarchy")
    print("   • Better contrast and readability")
def main():
    input_file = "/content/President Report_2 (1).xlsx"
    output_file = "enhanced_president_report.xlsx"
    try:
        format_excel_file(input_file, output_file)
    except FileNotFoundError:
        print(f"❌ Error: Could not find input file '{input_file}'")
        print("Please check the file path and try again.")
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        print("Please check your file and try again.")
if __name__ == "__main__":
    main()
