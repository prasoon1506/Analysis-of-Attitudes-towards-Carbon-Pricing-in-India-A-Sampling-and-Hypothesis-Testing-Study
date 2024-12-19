import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from collections import Counter
import scipy.stats as stats
import numpy as np
import io
st.set_page_config(page_title="Brand Price Analysis Report Generator",page_icon="📊",layout="wide")
st.markdown("""
    <style>
    .main {
        padding: 2rem;
    }
    .stAlert {
        padding: 1rem;
        margin: 1rem 0;
    }
    </style>
""", unsafe_allow_html=True)
st.title("📊 Brand Price Analysis Report Generator")
st.markdown("""Generate detailed brand price analysis reports from your Excel data.The report includes statistical analysis by region, brand, and month.""")
def safe_stats_calculation(data):
    if len(data) == 0:
        return {'Median': 0,'Mode': 0,'Mode Frequency': 0,'Q1': 0,'Q3': 0,'Min': 0,'Max': 0,'Skewness': 0}
    try:
        mode_calc = Counter(data).most_common(1)[0]
        mode_value, mode_freq = mode_calc
    except IndexError:
        mode_value, mode_freq = 0, 0
    try:
        skewness = float(stats.skew(data))
    except:
        skewness = 0
    return {
        'Median': np.median(data),
        'Mode': mode_value,
        'Mode Frequency': mode_freq,
        'Q1': np.percentile(data, 25),
        'Q3': np.percentile(data, 75),
        'Min': min(data),
        'Max': max(data),
        'Skewness': skewness
    }

def format_excel_report(df):
    wb = Workbook()
    wb.remove(wb.active)

    # Styles
    header_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
    section_font = Font(name='Calibri', size=12, bold=True, color="000000")
    regular_font = Font(name='Calibri', size=11)
    money_font = Font(name='Calibri', size=11, bold=True)
    subtitle_font = Font(name='Calibri', size=11, italic=True, color="666666")
    
    header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
    section_fill = PatternFill(start_color="DBE5F1", end_color="DBE5F1", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F8FBFF", end_color="F8FBFF", fill_type="solid")
    subtitle_fill = PatternFill(start_color="EDF3FA", end_color="EDF3FA", fill_type="solid")
    
    thin_border = Side(border_style="thin", color="000000")
    thick_border = Side(border_style="medium", color="2F75B5")
    header_border = Border(left=thin_border, right=thin_border, top=thick_border, bottom=thick_border)
    regular_border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

    # Progress bar for processing
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    total_regions = len(df['new_region1'].unique())
    regions_processed = 0

    for region in df['new_region1'].unique():
        status_text.text(f"Processing region: {region}")
        
        ws = wb.create_sheet(title=region)
        region_data = df.loc[df['new_region1'] == region].copy()
        region_data.loc[:, 'Month'] = pd.to_datetime(region_data['checkin date'], format='%d/%m/%Y').dt.month_name()
        
        months = sorted(region_data['Month'].unique())
        brands = sorted(region_data['Brand: Name'].unique())
        
        # Calculate statistics
        monthly_stats = {}
        for brand in brands:
            monthly_stats[brand] = {}
            for month in months:
                brand_month_data = region_data[
                    (region_data['Brand: Name'] == brand) & 
                    (region_data['Month'] == month)
                ]['Whole Sale Price'].values
                
                monthly_stats[brand][month] = safe_stats_calculation(brand_month_data)
                monthly_stats[brand][month]['Mean'] = np.mean(brand_month_data) if len(brand_month_data) > 0 else 0
                monthly_stats[brand][month]['Count'] = len(brand_month_data)
        
        # Format worksheet
        # [Previous formatting code remains the same...]
        # [Copy all the formatting code from the previous version]
        
        regions_processed += 1
        progress_bar.progress(regions_processed / total_regions)

    # Save to BytesIO object instead of file
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer

def main():
    # File upload section
    st.header("1. Upload Your Data")
    uploaded_file = st.file_uploader(
        "Upload your Excel file (must contain columns: new_region1, Brand: Name, checkin date, Whole Sale Price)",
        type=['xlsx', 'xls']
    )
    
    if uploaded_file is not None:
        try:
            # Load data
            with st.spinner('Reading your Excel file...'):
                df = pd.read_excel(uploaded_file)
            
            # Validate required columns
            required_columns = ['new_region1', 'Brand: Name', 'checkin date', 'Whole Sale Price']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                st.error(f"Missing required columns: {', '.join(missing_columns)}")
                return
            
            # Display data preview
            st.header("2. Data Preview")
            st.dataframe(df.head())
            
            # Display summary statistics
            st.header("3. Data Summary")
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Total Regions", len(df['new_region1'].unique()))
            with col2:
                st.metric("Total Brands", len(df['Brand: Name'].unique()))
            with col3:
                st.metric("Total Records", len(df))
            
            # Generate report button
            st.header("4. Generate Report")
            if st.button("Generate Excel Report", type="primary"):
                with st.spinner('Generating report...'):
                    # Generate the report
                    excel_buffer = format_excel_report(df)
                    
                    # Offer download
                    st.success("Report generated successfully!")
                    st.download_button(
                        label="📥 Download Excel Report",
                        data=excel_buffer,
                        file_name="brand_price_analysis_report.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
        
        except Exception as e:
            st.error(f"An error occurred: {str(e)}")
            st.markdown("Please make sure your Excel file is properly formatted and contains the required columns.")

if __name__ == "__main__":
    main()
