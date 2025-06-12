import streamlit as st
import openpyxl
from datetime import datetime
import calendar
import pandas as pd
import base64
from openpyxl.styles import (Font, Alignment, Border, Side, PatternFill, NamedStyle)
import streamlit as st
import openpyxl
from datetime import datetime
import calendar
import pandas as pd
import base64
from openpyxl.styles import (Font, Alignment, Border, Side, PatternFill, NamedStyle, Color)
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
def calculate_projected_usage(full_month_plan, input_date, month):
    _, total_days = calendar.monthrange(datetime.now().year, datetime.strptime(month, "%b").month)
    day = int(input_date.split()[0])
    projected_usage = (day / total_days) * full_month_plan
    return projected_usage
def calculate_actual_usage_percentage(actual_usage, full_month_plan):
    try:
        return (actual_usage / full_month_plan) * 100 if full_month_plan != 0 else 0
    except (TypeError, ZeroDivisionError):
        return 0
def calculate_pro_rata_deviation(actual_usage_percentage, input_date, month):
    _, total_days = calendar.monthrange(datetime.now().year, 
        datetime.strptime(month, "%b").month)
    day = int(input_date.split()[0])
    pro_rata_expectation = (day / total_days) * 100
    return actual_usage_percentage - pro_rata_expectation
def calculate_average_consumption(actual_usage, input_date):
    day = int(input_date.split()[0])
    try:
        return actual_usage / day if day != 0 else 0
    except (TypeError, ZeroDivisionError):
        return 0
def calculate_days_stock_available(current_stock, avg_consumption):
    try:
        return current_stock / avg_consumption if avg_consumption != 0 else 0
    except (TypeError, ZeroDivisionError):
        return 0
def style_excel(df, output_path):
    writer = pd.ExcelWriter(output_path, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='Bag Consumption Report')
    workbook = writer.book
    worksheet = writer.sheets['Bag Consumption Report']
    colors = {'dark_blue': '1E4C7B','light_blue': '4A90E2','header_bg': '2C3E50','alternate_row': 'F0F8FF','text_primary': '2C3E50','text_secondary': '34495E'}
    border = Border(left=Side(style='thin', color=Color(rgb=colors['light_blue'])),right=Side(style='thin', color=Color(rgb=colors['light_blue'])),top=Side(style='thin', color=Color(rgb=colors['light_blue'])),bottom=Side(style='thin', color=Color(rgb=colors['light_blue'])))
    header_fill = PatternFill(start_color=colors['header_bg'], end_color=colors['header_bg'],fill_type='solid')
    alternate_fill = PatternFill(start_color=colors['alternate_row'],end_color=colors['alternate_row'],fill_type='solid')
    for cell in worksheet[1]:
        cell.font = Font(bold=True,color='FFFFFF',name='Calibri',size=12)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        cell.border = border
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
            if cell.row % 2 == 0:
                cell.fill = alternate_fill
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        adjusted_width = min(max_length + 3, 50)
        worksheet.column_dimensions[column].width = adjusted_width
    worksheet.freeze_panes = worksheet['B2']
    worksheet.title = 'Bag Consumption Report'
    writer.close()
def filter_and_rename_columns(input_file, merge_file, user_date):
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active
    merge_wb = openpyxl.load_workbook(merge_file)
    merge_ws = merge_wb.active
    merge_data = {}
    for row in merge_ws.iter_rows(min_row=2):
        key = (str(row[0].value).strip() if row[0].value is not None else '', str(row[1].value).strip() if row[1].value is not None else '',str(row[2].value).strip() if row[2].value is not None else '')
        merge_data[key] = str(row[3].value).strip() if row[3].value is not None else ''
    columns_to_keep = [1, 2, 3, 4, 6, 8, 9]
    data_rows = []
    header = []
    input_day, input_month = user_date.split()
    _, total_days = calendar.monthrange(datetime.now().year, 
        datetime.strptime(input_month, "%b").month)
    for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row), 1):
        new_row = []
        row_values = []
        for idx, cell in enumerate(row, 1):
            if idx in columns_to_keep:
                new_row.append(cell.value)
                if idx <= 3:
                    row_values.append(str(cell.value).strip() if cell.value is not None else '')
        row_match_key = tuple(row_values)
        if row_num == 1:
            header = ["Plant Name","Brand Name", "Bag Name", "Opening Balance as on 01.06.2025","Tomonth Receipt",f"Actual Usage (Till {user_date})","Current available stock","Full Month Plan",f"Projected Usage (Till {user_date})",f"Actual Usage % (Till {user_date}) (Based on Planning)","Pro Rata Deviation","Average Consumption","No. of Days Stock Left (Based on Consumption)","No. of Days Stock Left (Based on Planning)"]
            continue
        full_month_plan = merge_data.get(row_match_key, "0")
        try:
            full_month_plan = float(full_month_plan)
        except ValueError:
            full_month_plan = 0
        new_row.append(full_month_plan)
        projected_usage = calculate_projected_usage(full_month_plan, user_date, input_month)
        new_row.append(int(projected_usage))
        actual_usage = new_row[5]
        opening_balance = new_row[3]
        current_stock = new_row[6]
        actual_usage_percentage = calculate_actual_usage_percentage(actual_usage, full_month_plan)
        new_row.append(int(actual_usage_percentage))
        pro_rata_deviation = calculate_pro_rata_deviation(actual_usage_percentage, user_date, input_month)
        new_row.append(int(pro_rata_deviation))
        average_consumption = calculate_average_consumption(actual_usage, user_date)
        new_row.append(int(average_consumption) if average_consumption is not None else 0)
        days_stock_tomonth_receipt = calculate_days_stock_available(current_stock, average_consumption)
        new_row.append(int(days_stock_tomonth_receipt))
        try:
            days_stock_planning = calculate_days_stock_available(opening_balance + full_month_plan - actual_usage,average_consumption)
        except (TypeError, ValueError):
            days_stock_planning = 0
        new_row.append(int(days_stock_planning))
        if isinstance(new_row[0], str) and len(new_row[0]) > 9:
            new_row[0] = new_row[0][9:]
        if isinstance(new_row[2], str) and len(new_row[2]) > 9:
            new_row[2] = new_row[2][9:]
        data_rows.append(new_row)
    df = pd.DataFrame(data_rows, columns=header)
    return df
def get_download_link(df):
    output_path = 'bag_report.xlsx'
    style_excel(df, output_path)
    with open(output_path, 'rb') as f:
        bytes = f.read()
    b64 = base64.b64encode(bytes).decode()
    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="bag_report.xlsx">Download Professional Excel Report</a>'
    return href
def main():
    st.set_page_config(page_title="Bag Report", layout="wide", page_icon="📊")
    st.markdown("""<style>.reportview-container {background-color: #f0f2f6;}.big-font {font-size:24px !important;color: #1E4C7B;font-weight: bold;text-transform: uppercase;letter-spacing: 1px;}.sub-font {font-size:18px !important;color: #4A90E2;font-style: italic;}</style>""", unsafe_allow_html=True)
    st.markdown('<h1 style="color:#1E4C7B; text-align:center; border-bottom: 3px solid #4A90E2; padding-bottom: 10px;">📊 Bag Consumption Report</h1>', unsafe_allow_html=True)
    col1, col2, col3 = st.columns(3)
    with col1:
        input_file = st.file_uploader("Upload Input Excel File", type=['xlsx', 'xls'], help="Select the input file for analysis")
    with col2:
        merge_file = st.file_uploader("Upload Merge Excel File", type=['xlsx', 'xls'], help="Select the merge file for additional data")
    with col3:
        user_date = st.text_input("Enter Date (e.g., 01 Jun)", value="01 Jun", help="Enter the date for report generation")
    if input_file and merge_file and user_date:
        try:
            with open("input_file.xlsx", "wb") as f:
                f.write(input_file.getbuffer())
            with open("merge_file.xlsx", "wb") as f:
                f.write(merge_file.getbuffer())
            df = filter_and_rename_columns("input_file.xlsx", "merge_file.xlsx", user_date)
            start_date = "01 Jun"
            st.markdown(f'<p class="big-font">Consumption Analysis</p>', unsafe_allow_html=True)
            st.markdown(f'<p class="sub-font">Period: {start_date} to {user_date}</p>', unsafe_allow_html=True)
            st.dataframe(df, use_container_width=True)
            st.markdown(get_download_link(df), unsafe_allow_html=True)
        except Exception as e:
            st.error(f"An error occurred: {str(e)}")
if __name__ == "__main__":
    main()
