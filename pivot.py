import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import streamlit as st
import io
import base64
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
def create_download_link(file, filename):
    b64 = base64.b64encode(file).decode()
    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}">Download Excel file</a>'
    return href
def create_navigator_excel(uploaded_file):
    file_data = io.BytesIO(uploaded_file.getvalue())
    wb = openpyxl.load_workbook(file_data)
    sheet_names = wb.sheetnames
    if "Dashboard" in wb.sheetnames:
        # If Dashboard already exists, remove it to recreate
        idx = wb.sheetnames.index("Dashboard")
        wb.remove(wb.worksheets[idx])
    dashboard = wb.create_sheet("Dashboard", 0)
    dashboard['A1'] = "SHEET NAVIGATOR"
    dashboard['A1'].font = Font(size=16, bold=True)
    dashboard.merge_cells('A1:D1')
    dashboard['A1'].alignment = Alignment(horizontal='center')
    dashboard['A3'] = "Select a sheet to view:"
    dashboard['A3'].font = Font(bold=True)
    selection_cell = 'B3'
    sheet_options = [name for name in sheet_names if name != "Dashboard"]
    sheet_list = ','.join([f'"{name}"' for name in sheet_options])
    dv = DataValidation(type="list", formula1=f"=\"{sheet_list}\"", allow_blank=False)
    dashboard.add_data_validation(dv)
    dv.add(selection_cell)
    dashboard[selection_cell] = sheet_options[0] if sheet_options else ""
    dashboard['A5'] = "NOTE: To view a sheet, select it from the dropdown above."
    dashboard['A5'].font = Font(italic=True, size=10)
    dashboard.merge_cells('A5:D5')
    dashboard['A7'] = "All Available Sheets:"
    dashboard['A7'].font = Font(bold=True)
    row = 8
    for i, sheet_name in enumerate(sheet_options):
        dashboard.cell(row=row, column=1, value=sheet_name)
        row += 1
    preview_row = row + 2
    dashboard.cell(row=preview_row, column=1, value="Data Preview (First 10 rows):")
    dashboard.cell(row=preview_row, column=1).font = Font(bold=True)
    dashboard.merge_cells(f'A{preview_row}:D{preview_row}')
    preview_row += 1
    dashboard.cell(row=preview_row, column=1, value="(Select a sheet to see data preview)")
    dashboard.cell(row=preview_row, column=1).alignment = Alignment(italic=True)
    dashboard.merge_cells(f'A{preview_row}:D{preview_row}')
    defined_name = DefinedName('SheetSelector', attr_text=f'Dashboard!${selection_cell}')
    wb.defined_names.append(defined_name)
    for col in range(1, 5):
        dashboard.column_dimensions[get_column_letter(col)].width = 20
    instruction_row = preview_row + 5
    dashboard.cell(row=instruction_row, column=1, value="INSTRUCTIONS:")
    dashboard.cell(row=instruction_row, column=1).font = Font(bold=True, size=12)
    dashboard.merge_cells(f'A{instruction_row}:D{instruction_row}')
    instruction_row += 1
    instructions = ["1. Select a sheet name from the dropdown above.","2. Press Alt+F8 to open the Macro dialog (we'll use Excel's built-in functionality).","3. Type 'ViewSheet' and click 'Create'.","4. Copy and paste this code (this is a one-time setup):","","Sub ViewSheet()","   Dim ws As Worksheet","   Dim selectedSheet As String","   selectedSheet = Range(\"SheetSelector\").Value","","   ' Hide all sheets except Dashboard","   For Each ws In ThisWorkbook.Worksheets","       If ws.Name <> \"Dashboard\" Then","           If ws.Name = selectedSheet Then","               ws.Visible = xlSheetVisible","           Else","               ws.Visible = xlSheetHidden","           End If","       End If","   Next ws","","   ' Activate the selected sheet","   ThisWorkbook.Sheets(selectedSheet).Activate","End Sub","","NOTE: This setup is needed only once per file. After setup, you can use","the macro directly or add a button to call it."]
    for i, line in enumerate(instructions):
        cell = dashboard.cell(row=instruction_row + i, column=1, value=line)
        dashboard.merge_cells(f'A{instruction_row + i}:D{instruction_row + i}')
        if "Sub ViewSheet" in line or line.startswith("   "):
            cell.font = Font(name="Courier New", size=9)
    for sheet_name in sheet_names:
        if sheet_name != "Dashboard":
            wb[sheet_name].sheet_state = 'hidden'
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
def create_no_macro_solution(uploaded_file):
    file_data = io.BytesIO(uploaded_file.getvalue())
    wb = openpyxl.load_workbook(file_data)
    sheet_names = wb.sheetnames
    xls = pd.ExcelFile(file_data)
    dashboard = wb.create_sheet("Dashboard", 0)
    dashboard['A1'] = "SHEET INDEX"
    dashboard['A1'].font = Font(size=16, bold=True)
    dashboard.merge_cells('A1:D1')
    dashboard['A1'].alignment = Alignment(horizontal='center')
    dashboard['A3'] = "Click on a sheet name below to navigate:"
    dashboard['A3'].font = Font(italic=True)
    dashboard.merge_cells('A3:D3')
    row = 5
    button_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    button_font = Font(color="FFFFFF", bold=True)
    for i, sheet_name in enumerate(sheet_names):
        if sheet_name != "Dashboard":
            cell = dashboard.cell(row=row, column=1, value=sheet_name)
            cell.font = button_font
            cell.fill = button_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
            cell.hyperlink = f"#{sheet_name}!A1"
            dashboard.cell(row=row, column=2, value="Preview:")
            try:
                sheet_data = pd.read_excel(xls, sheet_name=sheet_name)
                preview_data = sheet_data.iloc[:3, :3]
                for r_idx, row_data in enumerate(preview_data.values):
                    for c_idx, value in enumerate(row_data[:3]):
                        dashboard.cell(row=row+r_idx, column=3+c_idx, value=str(value)[:20])
            except:
                dashboard.cell(row=row, column=3, value="(Preview not available)")
            row += 4
    instruction_row = row + 2
    dashboard.cell(row=instruction_row, column=1, value="How to use:")
    dashboard.cell(row=instruction_row, column=1).font = Font(bold=True)
    instruction_row += 1
    instructions = ["1. Click on any sheet name to navigate to that sheet","2. To return to this Dashboard, use Excel's sheet tabs at the bottom","3. To hide sheets, right-click on a sheet tab and select 'Hide'","4. To unhide sheets, right-click on any visible sheet tab and select 'Unhide'"]
    for i, instruction in enumerate(instructions):
        dashboard.cell(row=instruction_row + i, column=1, value=instruction)
        dashboard.merge_cells(f'A{instruction_row + i}:D{instruction_row + i}')
    dashboard.column_dimensions['A'].width = 25
    dashboard.column_dimensions['B'].width = 15
    dashboard.column_dimensions['C'].width = 20
    dashboard.column_dimensions['D'].width = 20
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
def create_alternative_solution(uploaded_file):
    file_data = io.BytesIO(uploaded_file.getvalue())
    xls = pd.ExcelFile(file_data)
    sheet_names = xls.sheet_names
    wb_index = openpyxl.Workbook()
    index_sheet = wb_index.active
    index_sheet.title = "Sheet Index"
    index_sheet['A1'] = "SHEET INDEX"
    index_sheet['A1'].font = Font(size=16, bold=True)
    index_sheet.merge_cells('A1:C1')
    index_sheet['A1'].alignment = Alignment(horizontal='center')
    index_sheet['A3'] = "Sheet Name"
    index_sheet['B3'] = "Preview"
    index_sheet['C3'] = "Row Count"
    for col in ['A', 'B', 'C']:
        cell = index_sheet[f'{col}3']
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    row = 4
    for sheet_name in sheet_names:
        index_sheet.cell(row=row, column=1, value=sheet_name)
        try:
            df = pd.read_excel(xls, sheet_name=sheet_name)
            index_sheet.cell(row=row, column=3, value=len(df))
            if not df.empty:
                preview = str(df.iloc[0, 0]) if df.shape[1] > 0 else ""
                preview = (preview[:15] + '...') if len(str(preview)) > 15 else preview
                index_sheet.cell(row=row, column=2, value=preview)
        except:
            index_sheet.cell(row=row, column=2, value="(No preview)")
            index_sheet.cell(row=row, column=3, value="N/A")
        row += 1
    index_sheet.column_dimensions['A'].width = 25
    index_sheet.column_dimensions['B'].width = 20
    index_sheet.column_dimensions['C'].width = 15
    instruction_row = row + 2
    index_sheet.cell(row=instruction_row, column=1, value="INSTRUCTIONS FOR VIEWING SHEETS:")
    index_sheet.cell(row=instruction_row, column=1).font = Font(bold=True)
    index_sheet.merge_cells(f'A{instruction_row}:C{instruction_row}')
    instruction_row += 1
    instructions = ["1. This is an index of all sheets in the original workbook","2. To view individual sheets, they must be individually unhidden in Excel","3. To unhide a sheet: Right-click on any sheet tab → Select 'Unhide' → Choose sheet","4. To hide a sheet: Right-click on the sheet tab → Select 'Hide'"]
    for i, instruction in enumerate(instructions):
        index_sheet.cell(row=instruction_row + i, column=1, value=instruction)
        index_sheet.merge_cells(f'A{instruction_row + i}:C{instruction_row + i}')
    output = io.BytesIO()
    wb_index.save(output)
    output.seek(0)
    return output
def main():
    st.title("Excel Sheet Navigator Creator")
    st.write("""## Upload your Excel fileThis tool will create a new Excel file with a navigation dashboard that allows you to:- See all available sheets in your workbook- View data previews from each sheet- Navigate between sheets easily""")
    uploaded_file = st.file_uploader("Choose an Excel file", type=["xlsx", "xls"])
    if uploaded_file is not None:
        file_details = {"Filename": uploaded_file.name, "File size": f"{uploaded_file.size/1024:.1f} KB"}
        st.write(file_details)
        st.write("### Select your preferred solution:")
        solution_type = st.radio("Choose a navigation approach:",["Simple Navigation (No Macros)","Advanced Navigation (Requires One-Time Macro Setup)","Basic Index Sheet"])
        if st.button("Generate Excel File"):
            try:
                with st.spinner("Creating Excel file..."):
                    if solution_type == "Simple Navigation (No Macros)":
                        excel_file = create_no_macro_solution(uploaded_file)
                        download_filename = f"Navigator_{uploaded_file.name}"
                        solution_description = """
                        ### Simple Navigation- Click on sheet names to navigate directly to those sheets- Use Excel's sheet tabs to return to the Dashboard- No macros required, but sheets won't be automatically hidden"""
                    elif solution_type == "Advanced Navigation (Requires One-Time Macro Setup)":
                        excel_file = create_navigator_excel(uploaded_file)
                        download_filename = f"MacroNavigator_{uploaded_file.name}"
                        solution_description = """### Advanced Navigation (One-Time Macro Setup)- Follow the instructions in the Dashboard sheet to set up the macro- Select sheets from the dropdown and run the macro (Alt+F8 → ViewSheet → Run)- This solution provides full hide/show functionality"""
                    else:
                        excel_file = create_alternative_solution(uploaded_file)
                        download_filename = f"Index_{uploaded_file.name}"
                        solution_description = """### Basic Index Sheet- Simple index of all sheets in your workbook- Use Excel's built-in unhide functionality to view sheets- Minimal approach with no hyperlinks or macros"""
                st.markdown(create_download_link(excel_file.getvalue(), download_filename),unsafe_allow_html=True)
                st.success("✅ Your Excel file has been created!")
                st.markdown(solution_description)
            except Exception as e:
                st.error(f"An error occurred: {str(e)}")
                st.error("Please make sure your Excel file is not corrupted and try again.")
if __name__ == "__main__":
    main()
