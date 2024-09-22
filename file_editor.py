import streamlit as st
import pandas as pd
import numpy as np
import base64
from io import BytesIO
import openpyxl

# Set page config
st.set_page_config(page_title="Excel Editor", layout="wide")

# Custom CSS for styling
st.markdown("""
<style>
    .main .block-container {
        padding-top: 2rem;
    }
    h1 {
        color: #2c3e50;
        text-align: center;
        margin-bottom: 2rem;
    }
    .stButton > button {
        width: 100%;
    }
    .excel-table {
        border-collapse: collapse;
        width: 100%;
    }
    .excel-table th, .excel-table td {
        border: 1px solid #ddd;
        padding: 8px;
        text-align: left;
    }
    .excel-table tr:nth-child(even) {
        background-color: #f2f2f2;
    }
    .excel-table th {
        padding-top: 12px;
        padding-bottom: 12px;
        background-color: #4CAF50;
        color: white;
    }
</style>
""", unsafe_allow_html=True)

# Title
st.title("Interactive Excel Editor")

# Function to create HTML representation of Excel structure (first 5 rows)
def create_excel_structure_html(sheet, max_rows=5):
    html = "<table class='excel-table'>"
    merged_cells = sheet.merged_cells.ranges

    for idx, row in enumerate(sheet.iter_rows(max_row=max_rows)):
        html += "<tr>"
        for cell in row:
            merged = False
            for merged_range in merged_cells:
                if cell.coordinate in merged_range:
                    if cell.coordinate == merged_range.start_cell.coordinate:
                        rowspan = min(merged_range.max_row - merged_range.min_row + 1, max_rows - idx)
                        colspan = merged_range.max_col - merged_range.min_col + 1
                        html += f"<td rowspan='{rowspan}' colspan='{colspan}'>{cell.value}</td>"
                    merged = True
                    break
            if not merged:
                html += f"<td>{cell.value}</td>"
        html += "</tr>"
    html += "</table>"
    return html

# Function to get merged column groups
def get_merged_column_groups(sheet):
    merged_groups = {}
    for merged_range in sheet.merged_cells.ranges:
        if merged_range.min_row == 1:  # Only consider merged cells in the first row (header)
            main_col = sheet.cell(1, merged_range.min_col).value
            merged_groups[main_col] = list(range(merged_range.min_col, merged_range.max_col + 1))
    return merged_groups

# File uploader
uploaded_file = st.file_uploader("Choose an Excel file", type="xlsx")

if uploaded_file is not None:
    # Read Excel file
    excel_file = openpyxl.load_workbook(uploaded_file)
    sheet = excel_file.active

    # Display original Excel structure (first 5 rows)
    st.subheader("Original Excel Structure (First 5 Rows)")
    excel_html = create_excel_structure_html(sheet, max_rows=5)
    st.markdown(excel_html, unsafe_allow_html=True)

    # Get merged column groups
    merged_groups = get_merged_column_groups(sheet)

    # Create a list of column headers, considering merged cells
    column_headers = []
    column_indices = {}  # To store the column indices for each header
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, col).value
        if cell_value is not None:
            column_headers.append(cell_value)
            if cell_value not in column_indices:
                column_indices[cell_value] = []
            column_indices[cell_value].append(col - 1)  # pandas uses 0-based index
        else:
            # If the cell is empty, it's part of a merged cell, so use the previous header
            prev_header = column_headers[-1]
            column_headers.append(prev_header)
            column_indices[prev_header].append(col - 1)

    # Read as pandas DataFrame using the correct column headers
    df = pd.read_excel(uploaded_file, header=None, names=column_headers)
    df = df.iloc[1:]  # Remove the first row as it's now our header

    # Column selection for deletion
    st.subheader("Select columns to delete")
    all_columns = list(set(column_headers))  # Use set to get unique column names
    cols_to_delete = st.multiselect("Choose columns to remove", all_columns)
    
    if cols_to_delete:
        columns_to_remove = []
        for col in cols_to_delete:
            columns_to_remove.extend(column_indices[col])
        
        df = df.drop(df.columns[columns_to_remove], axis=1)
        st.success(f"Deleted columns: {', '.join(cols_to_delete)}")

    # Row deletion
    st.subheader("Delete rows")
    num_rows = st.number_input("Enter the number of rows to delete from the start", min_value=0, max_value=len(df)-1, value=0)
    
    if num_rows > 0:
        df = df.iloc[num_rows:]
        st.success(f"Deleted first {num_rows} rows")
    
    # Display editable dataframe
    st.subheader("Edit Data")
    st.write("You can edit individual cell values directly in the table below:")
    
    # Replace NaN values with None and convert dataframe to a dictionary
    df_dict = df.where(pd.notnull(df), None).to_dict('records')
    
    # Use st.data_editor with the processed dictionary
    edited_data = st.data_editor(df_dict)
    
    # Convert edited data back to dataframe
    edited_df = pd.DataFrame(edited_data)
    
    # Display edited data
    st.subheader("Edited Data")
    st.dataframe(edited_df)
    
    # Download button
    def get_excel_download_link(df):
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
        excel_data = output.getvalue()
        b64 = base64.b64encode(excel_data).decode()
        return f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="edited_file.xlsx">Download Edited Excel File</a>'
    
    st.markdown(get_excel_download_link(edited_df), unsafe_allow_html=True)

else:
    st.info("Please upload an Excel file to begin editing.")

# Add some final instructions
st.markdown("""
---
### Instructions:
1. Upload an Excel file using the file uploader at the top of the page.
2. View the original Excel structure (first 5 rows), including merged cells.
3. Select columns to delete. For merged headers, selecting any part will delete the entire merged group.
4. Specify the number of rows to delete from the start, if any.
5. Edit individual cell values directly in the editable table.
6. Review your changes in the "Edited Data" section.
7. Download the edited Excel file using the download link provided.

Enjoy editing your Excel files with ease!
""")
