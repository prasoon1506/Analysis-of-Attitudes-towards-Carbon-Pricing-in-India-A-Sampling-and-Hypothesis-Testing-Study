import streamlit as st
import openpyxl
from datetime import datetime
import calendar
import pandas as pd
import base64
from openpyxl.styles import (Font, Alignment, Border, Side, PatternFill, NamedStyle, Color)
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
def calculate_projected_usage(full_month_plan, input_date, month):
    _, total_days = calendar.monthrange(datetime.now().year, datetime.strptime(month, "%b").month)
    day = int(input_date.split()[0])
    projected_usage = (day / total_days) * full_month_plan
    return projected_usage
def calculate_actual_usage_percentage(actual_usage, full_month_plan):
    try:
        return (actual_usage / full_month_plan) * 100 if full_month_plan != 0 else 0
    except (TypeError, ZeroDivisionError):
        return 0
def calculate_pro_rata_deviation(actual_usage_percentage, input_date, month):
    _, total_days = calendar.monthrange(datetime.now().year, 
        datetime.strptime(month, "%b").month)
    day = int(input_date.split()[0])
    pro_rata_expectation = (day / total_days) * 100
    return actual_usage_percentage - pro_rata_expectation
def calculate_average_consumption(actual_usage, input_date):
    day = int(input_date.split()[0])
    try:
        return actual_usage / day if day != 0 else 0
    except (TypeError, ZeroDivisionError):
        return 0
def calculate_days_stock_available(current_stock, avg_consumption):
    try:
        return current_stock / avg_consumption if avg_consumption != 0 else 0
    except (TypeError, ZeroDivisionError):
        return 0
def write_formulas_to_excel(df, output_path, user_date):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Bag Consumption Report'
    input_day, input_month = user_date.split()
    day = int(input_day)
    _, total_days = calendar.monthrange(datetime.now().year, datetime.strptime(input_month, "%b").month)
    headers = df.columns.tolist()
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF', name='Aptos Narrow', size=12)
    for row_idx, (index, row) in enumerate(df.iterrows(), 2):
        for col in range(1, 9):
            cell = ws.cell(row=row_idx, column=col, value=row.iloc[col-1])
            cell.font = Font(name='Aptos Narrow', size=11)
        projected_formula = f"=ROUND(({day}/{total_days})*H{row_idx},0)"
        cell = ws.cell(row=row_idx, column=9, value=projected_formula)
        cell.font = Font(name='Aptos Narrow', size=11)
        usage_percent_formula = f"=IF(H{row_idx}=0,0,(F{row_idx}/H{row_idx}))"
        cell = ws.cell(row=row_idx, column=10, value=usage_percent_formula)
        cell.font = Font(name='Aptos Narrow', size=11)
        pro_rata_formula = f"=J{row_idx}-({day}/{total_days})"
        cell = ws.cell(row=row_idx, column=11, value=pro_rata_formula)
        cell.font = Font(name='Aptos Narrow', size=11)
        avg_consumption_formula = f"=ROUND(IF({day}=0,0,F{row_idx}/{day}),0)"
        cell = ws.cell(row=row_idx, column=12, value=avg_consumption_formula)
        cell.font = Font(name='Aptos Narrow', size=11)
        days_stock_consumption_formula = f"=ROUND(IF(L{row_idx}=0,0,G{row_idx}/L{row_idx}),0)"
        cell = ws.cell(row=row_idx, column=13, value=days_stock_consumption_formula)
        cell.font = Font(name='Aptos Narrow', size=11)
        days_stock_planning_formula = f"=ROUND(IF(L{row_idx}=0,0,(D{row_idx}+H{row_idx}-F{row_idx})/L{row_idx}),0)"
        cell = ws.cell(row=row_idx, column=14, value=days_stock_planning_formula)
        cell.font = Font(name='Aptos Narrow', size=11)
    style_excel_with_formulas(wb, ws)
    wb.save(output_path)
def style_excel_with_formulas(workbook, worksheet):
    colors = {'dark_blue': '1E4C7B','light_blue': '4A90E2','header_bg': '2C3E50','alternate_row': 'F0F8FF','text_primary': '2C3E50','text_secondary': '34495E'}
    border = Border(left=Side(style='thin', color=Color(rgb=colors['light_blue'])),right=Side(style='thin', color=Color(rgb=colors['light_blue'])),top=Side(style='thin', color=Color(rgb=colors['light_blue'])),bottom=Side(style='thin', color=Color(rgb=colors['light_blue']))) 
    header_fill = PatternFill(start_color=colors['header_bg'],end_color=colors['header_bg'],fill_type='solid')
    alternate_fill = PatternFill(start_color=colors['alternate_row'],end_color=colors['alternate_row'],fill_type='solid')
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color='FFFFFF', name='Aptos Narrow', size=12)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    for row in worksheet.iter_rows(min_row=2):
        for col_idx, cell in enumerate(row, 1):
            if cell.font.name != 'Aptos Narrow':
                cell.font = Font(name='Aptos Narrow', size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
            if cell.row % 2 == 0:
                cell.fill = alternate_fill
            if col_idx in [4, 5, 6, 7, 8, 9, 12, 13, 14]:  # All number columns without decimals
                cell.number_format = '0'
            elif col_idx in [10, 11]:  # Percentage columns (Actual Usage % and Pro Rata Deviation)
                cell.number_format = '0%'
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.row == 1:
                    lines = str(cell.value).split('\n') if cell.value else ['']
                    cell_length = max(len(line) for line in lines)
                else:
                    cell_length = len(str(cell.value)) if cell.value else 0
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        worksheet.column_dimensions[column].width = adjusted_width
    worksheet.row_dimensions[1].height = 60
    worksheet.freeze_panes = worksheet['B2']
def style_excel(df, output_path):
    writer = pd.ExcelWriter(output_path, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='Bag Consumption Report')
    workbook = writer.book
    worksheet = writer.sheets['Bag Consumption Report']
    colors = {'dark_blue': '1E4C7B','light_blue': '4A90E2','header_bg': '2C3E50','alternate_row': 'F0F8FF','text_primary': '2C3E50','text_secondary': '34495E'}
    border = Border(left=Side(style='thin', color=Color(rgb=colors['light_blue'])),right=Side(style='thin', color=Color(rgb=colors['light_blue'])),top=Side(style='thin', color=Color(rgb=colors['light_blue'])),bottom=Side(style='thin', color=Color(rgb=colors['light_blue'])))
    header_fill = PatternFill(start_color=colors['header_bg'],end_color=colors['header_bg'],fill_type='solid')
    alternate_fill = PatternFill(start_color=colors['alternate_row'],end_color=colors['alternate_row'],fill_type='solid')
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color='FFFFFF', name='Aptos Narrow', size=12)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    for row in worksheet.iter_rows(min_row=2):
        for col_idx, cell in enumerate(row, 1):
            cell.font = Font(name='Aptos Narrow', size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
            if cell.row % 2 == 0:
                cell.fill = alternate_fill
            if col_idx in [4, 5, 6, 7, 9, 12, 13, 14]:
                cell.number_format = '0'
            elif col_idx in [10, 11]:
                cell.number_format = '0%'
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.row == 1:
                    lines = str(cell.value).split('\n') if cell.value else ['']
                    cell_length = max(len(line) for line in lines)
                else:
                    cell_length = len(str(cell.value)) if cell.value else 0
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        worksheet.column_dimensions[column].width = adjusted_width
    worksheet.row_dimensions[1].height = 60
    worksheet.freeze_panes = worksheet['B2']
    worksheet.title = 'Bag Consumption Report'
    writer.close()
def filter_and_rename_columns(input_file, merge_file, user_date):
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active
    merge_wb = openpyxl.load_workbook(merge_file)
    merge_ws = merge_wb.active
    merge_data = {}
    for row in merge_ws.iter_rows(min_row=2):
        key = (str(row[0].value).strip() if row[0].value is not None else '', str(row[1].value).strip() if row[1].value is not None else '',str(row[2].value).strip() if row[2].value is not None else '')
        merge_data[key] = str(row[3].value).strip() if row[3].value is not None else ''
    columns_to_keep = [1, 2, 3, 4, 6, 8, 9]
    data_rows = []
    header = []
    input_day, input_month = user_date.split()
    _, total_days = calendar.monthrange(datetime.now().year, 
        datetime.strptime(input_month, "%b").month)
    for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row), 1):
        new_row = []
        row_values = []
        for idx, cell in enumerate(row, 1):
            if idx in columns_to_keep:
                new_row.append(cell.value)
                if idx <= 3:
                    row_values.append(str(cell.value).strip() if cell.value is not None else '')
        row_match_key = tuple(row_values)
        if row_num == 1:
            header = ["Plant\nName","Brand\nName", "Bag\nName","Opening Balance\nas on 01.09.2025","Tomonth\nReceipt",f"Actual Usage\n(Till {user_date})","Current\navailable stock","Full Month\nPlan",f"Projected Usage\n(Till {user_date})",f"Actual Usage %\n(Till {user_date})\n(Based on Planning)","Pro Rata\nDeviation","Average\nConsumption","No. of Days\nStock Left\n(Based on Consumption)","No. of Days\nStock Left\n(Based on Planning)"]
            continue
        full_month_plan = merge_data.get(row_match_key, "0")
        try:
            full_month_plan = float(full_month_plan)
        except ValueError:
            full_month_plan = 0
        new_row.append(full_month_plan)
        projected_usage = calculate_projected_usage(full_month_plan, user_date, input_month)
        new_row.append(int(projected_usage))
        actual_usage = new_row[5]
        opening_balance = new_row[3]
        current_stock = new_row[6]
        actual_usage_percentage = calculate_actual_usage_percentage(actual_usage, full_month_plan)
        new_row.append(int(actual_usage_percentage))
        pro_rata_deviation = calculate_pro_rata_deviation(actual_usage_percentage, user_date, input_month)
        new_row.append(int(pro_rata_deviation))
        average_consumption = calculate_average_consumption(actual_usage, user_date)
        new_row.append(int(average_consumption) if average_consumption is not None else 0)
        days_stock_tomonth_receipt = calculate_days_stock_available(current_stock, average_consumption)
        new_row.append(int(days_stock_tomonth_receipt))
        try:
            days_stock_planning = calculate_days_stock_available(opening_balance + full_month_plan - actual_usage,average_consumption)
        except (TypeError, ValueError):
            days_stock_planning = 0
        new_row.append(int(days_stock_planning))
        if isinstance(new_row[0], str) and len(new_row[0]) > 9:
            new_row[0] = new_row[0][9:]
        if isinstance(new_row[2], str) and len(new_row[2]) > 9:
            new_row[2] = new_row[2][9:]
        if isinstance(new_row[0], str) and new_row[0].lower().strip() == "totals":
            continue
        data_rows.append(new_row)
    df = pd.DataFrame(data_rows, columns=header)
    return df
def get_download_link(df, user_date, use_formulas=True):
    output_path = 'bag_report.xlsx'
    if use_formulas:
        write_formulas_to_excel(df, output_path, user_date)
    else:
        style_excel(df, output_path)
    with open(output_path, 'rb') as f:
        bytes = f.read()
    b64 = base64.b64encode(bytes).decode()
    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="bag_report.xlsx">Download Professional Excel Report with Formulas</a>'
    return href
def create_dashboard_charts(df):
    if df.empty:
        return None, None, None
    fig1 = go.Figure()
    current_stock = df.iloc[:, 6].astype(float)
    actual_usage = df.iloc[:, 5].astype(float)
    plant_names = df.iloc[:, 0].astype(str)
    fig1.add_trace(go.Bar(name='Current Stock',x=plant_names[:10],y=current_stock[:10],marker_color='rgba(74, 144, 226, 0.8)',hovertemplate='<b>%{x}</b><br>Current Stock: %{y}<extra></extra>'))
    fig1.add_trace(go.Bar(name='Actual Usage',x=plant_names[:10],y=actual_usage[:10],marker_color='rgba(255, 99, 132, 0.8)',hovertemplate='<b>%{x}</b><br>Actual Usage: %{y}<extra></extra>'))
    fig1.update_layout(title={'text': '📊 Stock vs Usage Analysis','x': 0.5,'font': {'size': 20, 'color': '#1E4C7B'}},barmode='group',xaxis_title='Plant Names',yaxis_title='Quantity',plot_bgcolor='rgba(248, 249, 250, 0.8)',paper_bgcolor='white',font=dict(family="Arial", size=12),height=400,xaxis={'tickangle': 45})
    avg_usage_percent = df.iloc[:, 9].astype(float).mean()
    fig2 = go.Figure(go.Indicator(mode = "gauge+number+delta",value = avg_usage_percent,domain = {'x': [0, 1], 'y': [0, 1]},title = {'text': "Overall Usage Efficiency %", 'font': {'size': 20, 'color': '#1E4C7B'}},delta = {'reference': 100},
        gauge = {'axis': {'range': [None, 100]},'bar': {'color': "#1E4C7B"},'steps': [{'range': [0, 50], 'color': "#FFE4E1"},{'range': [50, 80], 'color': "#B0E0E6"},{'range': [80, 100], 'color': "#98FB98"}],'threshold': {'line': {'color': "red", 'width': 4},'thickness': 0.75,'value': 90}}))
    fig2.update_layout(height=400,paper_bgcolor='white',font=dict(family="Arial", size=12))
    days_stock_consumption = df.iloc[:, 12].astype(float)
    fig3 = go.Figure(data=[go.Histogram(x=days_stock_consumption,nbinsx=20,marker_color='rgba(74, 144, 226, 0.7)',hovertemplate='Days: %{x}<br>Count: %{y}<extra></extra>')])
    fig3.update_layout(title={'text': '📈 Days Stock Left Distribution','x': 0.5,'font': {'size': 20, 'color': '#1E4C7B'}},xaxis_title='Days Stock Left',yaxis_title='Number of Items',plot_bgcolor='rgba(248, 249, 250, 0.8)',paper_bgcolor='white',font=dict(family="Arial", size=12),height=400)
    return fig1, fig2, fig3
def main():
    st.set_page_config(page_title="Advanced Bag Report Analytics", layout="wide", page_icon="📊",initial_sidebar_state="expanded")
    st.markdown("""<style>@import url('https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;600;700&display=swap');.main-header {background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);padding: 2rem;border-radius: 15px;margin-bottom: 2rem;text-align: center;box-shadow: 0 10px 30px rgba(0,0,0,0.1);animation: slideDown 0.8s ease-out;}
    .main-title {color: white;font-family: 'Poppins', sans-serif;font-size: 3rem;font-weight: 700;margin: 0;text-shadow: 2px 2px 4px rgba(0,0,0,0.3);}
    .main-subtitle {color: rgba(255,255,255,0.9);font-family: 'Poppins', sans-serif;font-size: 1.2rem;font-weight: 300;margin: 0.5rem 0 0 0;}
    .upload-card {background: white;padding: 1.5rem;border-radius: 15px;box-shadow: 0 5px 15px rgba(0,0,0,0.08);border-left: 4px solid #667eea;transition: transform 0.3s ease, box-shadow 0.3s ease;margin-bottom: 1rem;}
    .upload-card:hover {transform: translateY(-2px);box-shadow: 0 8px 25px rgba(0,0,0,0.12);}
    .metric-card {
        background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);
        padding: 1.5rem;
        border-radius: 15px;
        color: white;
        text-align: center;
        margin-bottom: 1rem;
        box-shadow: 0 5px 15px rgba(0,0,0,0.1);
    }
    
    .metric-title {
        font-family: 'Poppins', sans-serif;
        font-size: 0.9rem;
        font-weight: 400;
        opacity: 0.9;
        margin-bottom: 0.5rem;
    }
    
    .metric-value {
        font-family: 'Poppins', sans-serif;
        font-size: 2rem;
        font-weight: 700;
        margin: 0;
    }
    
    .feature-toggle {
        background: linear-gradient(135deg, #4facfe 0%, #00f2fe 100%);
        padding: 1rem;
        border-radius: 10px;
        margin: 1rem 0;
    }
    
    .status-good {
        color: #28a745;
        font-weight: bold;
    }
    
    .status-warning {
        color: #ffc107;
        font-weight: bold;
    }
    
    .status-danger {
        color: #dc3545;
        font-weight: bold;
    }
    
    .download-button {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        padding: 1rem 2rem;
        border-radius: 25px;
        text-decoration: none;
        font-weight: bold;
        display: inline-block;
        margin: 1rem 0;
        box-shadow: 0 5px 15px rgba(102, 126, 234, 0.4);
        transition: all 0.3s ease;
    }
    
    .download-button:hover {
        transform: translateY(-2px);
        box-shadow: 0 7px 20px rgba(102, 126, 234, 0.6);
        color: white;
        text-decoration: none;
    }
    
    .formula-section {
        background: linear-gradient(135deg, #fa709a 0%, #fee140 100%);
        padding: 1.5rem;
        border-radius: 15px;
        margin-top: 2rem;
        color: white;
    }
    
    .stSelectbox > div > div {
        border-radius: 10px;
    }
    
    .stTextInput > div > div {
        border-radius: 10px;
    }
    
    .stFileUploader > div {
        border-radius: 10px;
    }
    
    @keyframes slideDown {
        from { opacity: 0; transform: translateY(-30px); }
        to { opacity: 1; transform: translateY(0); }
    }
    
    .sidebar .sidebar-content {
        background: linear-gradient(180deg, #667eea 0%, #764ba2 100%);
    }
    
    .reportview-container .main .block-container {
        padding-top: 2rem;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Main header
    st.markdown("""
    <div class="main-header">
        <h1 class="main-title">🚀 Advanced Bag Report Analytics</h1>
        <p class="main-subtitle">Professional consumption analysis with real-time insights</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Sidebar for controls
    with st.sidebar:
        st.markdown("## ⚙️ Configuration Panel")
        
        # File upload section
        st.markdown("### 📁 Data Sources")
        input_file = st.file_uploader(
            "📊 Input Excel File", 
            type=['xlsx', 'xls'], 
            help="Upload your main consumption data file"
        )
        
        merge_file = st.file_uploader(
            "🔗 Merge Excel File", 
            type=['xlsx', 'xls'], 
            help="Upload the file containing additional planning data"
        )
        
        st.markdown("### 📅 Analysis Parameters")
        user_date = st.text_input(
            "📆 Report Date", 
            value="01 Nov", 
            help="Enter date in format: DD MMM (e.g., 15 Nov)"
        )
        
        st.markdown("### ⚡ Advanced Options")
        use_formulas = st.toggle(
            "Include Excel Formulas", 
            value=True,
            help="Toggle to include/exclude Excel formulas in the report"
        )
        
        if use_formulas:
            st.success("✅ Formulas will be embedded")
        else:
            st.info("📊 Values only mode")
    
    # Main content area
    if input_file and merge_file and user_date:
        try:
            # Save uploaded files
            with open("input_file.xlsx", "wb") as f:
                f.write(input_file.getbuffer())
            with open("merge_file.xlsx", "wb") as f:
                f.write(merge_file.getbuffer())
            
            # Process data
            with st.spinner('🔄 Processing your data...'):
                df = filter_and_rename_columns("input_file.xlsx", "merge_file.xlsx", user_date)
            
            # Success message
            st.success('✅ Data processed successfully!')
            
            # Key metrics section
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                total_items = len(df)
                st.markdown(f"""
                <div class="metric-card">
                    <div class="metric-title">Total Items</div>
                    <div class="metric-value">{total_items}</div>
                </div>
                """, unsafe_allow_html=True)
            
            with col2:
                total_stock = df.iloc[:, 6].astype(float).sum()
                st.markdown(f"""
                <div class="metric-card" style="background: linear-gradient(135deg, #4facfe 0%, #00f2fe 100%);">
                    <div class="metric-title">Total Current Stock</div>
                    <div class="metric-value">{total_stock:,.0f}</div>
                </div>
                """, unsafe_allow_html=True)
            
            with col3:
                total_usage = df.iloc[:, 5].astype(float).sum()
                st.markdown(f"""
                <div class="metric-card" style="background: linear-gradient(135deg, #fa709a 0%, #fee140 100%);">
                    <div class="metric-title">Total Usage</div>
                    <div class="metric-value">{total_usage:,.0f}</div>
                </div>
                """, unsafe_allow_html=True)
            
            with col4:
                avg_efficiency = df.iloc[:, 9].astype(float).mean()
                st.markdown(f"""
                <div class="metric-card" style="background: linear-gradient(135deg, #a8edea 0%, #fed6e3 100%);">
                    <div class="metric-title">Avg Efficiency</div>
                    <div class="metric-value">{avg_efficiency:.1f}%</div>
                </div>
                """, unsafe_allow_html=True)
            
            # Dashboard Charts Section
            st.markdown("## 📈 Interactive Analytics Dashboard")
            
            chart1, chart2, chart3 = create_dashboard_charts(df)
            
            if chart1 and chart2 and chart3:
                tab1, tab2, tab3 = st.tabs(["📊 Stock Analysis", "⚡ Efficiency Gauge", "📈 Distribution"])
                
                with tab1:
                    st.plotly_chart(chart1, use_container_width=True)
                    
                with tab2:
                    col1, col2 = st.columns([1, 1])
                    with col1:
                        st.plotly_chart(chart2, use_container_width=True)
                    with col2:
                        st.markdown("### 🎯 Performance Insights")
                        if avg_efficiency >= 80:
                            st.markdown('<p class="status-good">🟢 Excellent performance! Usage is well-aligned with planning.</p>', unsafe_allow_html=True)
                        elif avg_efficiency >= 60:
                            st.markdown('<p class="status-warning">🟡 Good performance with room for improvement.</p>', unsafe_allow_html=True)
                        else:
                            st.markdown('<p class="status-danger">🔴 Performance needs attention. Consider reviewing planning.</p>', unsafe_allow_html=True)
                        
                        # Additional metrics
                        low_stock_items = len(df[df.iloc[:, 12].astype(float) < 7])
                        st.metric("Items with <7 days stock", low_stock_items)
                        
                        high_usage_items = len(df[df.iloc[:, 9].astype(float) > 100])
                        st.metric("Over-consuming items", high_usage_items)
                
                with tab3:
                    st.plotly_chart(chart3, use_container_width=True)
            
            # Data table section
            st.markdown("## 📋 Detailed Report")
            
            # Filter options
            col1, col2 = st.columns(2)
            with col1:
                filter_option = st.selectbox(
                    "🔍 Filter by status:",
                    ["All Items", "Low Stock (<7 days)", "Over-consuming (>100%)", "Under-consuming (<50%)"]
                )
            
            with col2:
                sort_option = st.selectbox(
                    "📊 Sort by:",
                    ["Plant Name", "Current Stock", "Usage %", "Days Left"]
                )
            
            # Apply filters
            filtered_df = df.copy()
            if filter_option == "Low Stock (<7 days)":
                filtered_df = df[df.iloc[:, 12].astype(float) < 7]
            elif filter_option == "Over-consuming (>100%)":
                filtered_df = df[df.iloc[:, 9].astype(float) > 100]
            elif filter_option == "Under-consuming (<50%)":
                filtered_df = df[df.iloc[:, 9].astype(float) < 50]
            
            # Apply sorting
            if sort_option == "Current Stock":
                filtered_df = filtered_df.sort_values(filtered_df.columns[6], ascending=False)
            elif sort_option == "Usage %":
                filtered_df = filtered_df.sort_values(filtered_df.columns[9], ascending=False)
            elif sort_option == "Days Left":
                filtered_df = filtered_df.sort_values(filtered_df.columns[12], ascending=True)
            
            # Display filtered data
            st.dataframe(
                filtered_df,
                use_container_width=True,
                height=400
            )
            
            st.markdown(f"**Showing {len(filtered_df)} of {len(df)} items**")
            
            # Download section
            st.markdown("## 📥 Export Report")
            
            start_date = "01 Nov"
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown(f"""
                <div class="upload-card">
                    <h4>📊 Analysis Period</h4>
                    <p><strong>From:</strong> {start_date}</p>
                    <p><strong>To:</strong> {user_date}</p>
                    <p><strong>Formula Mode:</strong> {'Enabled' if use_formulas else 'Disabled'}</p>
                </div>
                """, unsafe_allow_html=True)
            
            with col2:
                download_link = get_download_link(df, user_date, use_formulas)
                st.markdown(f"""
                <div class="upload-card">
                    <h4>💾 Download Options</h4>
                    {download_link}
                </div>
                """, unsafe_allow_html=True)
            
            # Formula documentation (if enabled)
            if use_formulas:
                with st.expander("🧮 Formula Documentation", expanded=False):
                    st.markdown("""
                    <div class="formula-section">
                        <h3>📚 Excel Formula Reference</h3>
                        <p>The following formulas are embedded in your Excel report:</p>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    formula_data = {
                        "Column": [
                            "Projected Usage",
                            "Actual Usage %", 
                            "Pro Rata Deviation",
                            "Average Consumption",
                            "Days Stock Left (Consumption)",
                            "Days Stock Left (Planning)"
                        ],
                        "Formula": [
                            "=ROUND((Day/Total_Days) * Full_Month_Plan, 0)",
                            "=IF(Full_Month_Plan=0, 0, Actual_Usage/Full_Month_Plan)",
                            "=Actual_Usage% - (Day/Total_Days)",
                            "=ROUND(IF(Day=0, 0, Actual_Usage/Day), 0)",
                            "=ROUND(IF(Avg_Consumption=0, 0, Current_Stock/Avg_Consumption), 0)",
                            "=ROUND(IF(Avg_Consumption=0, 0, (Opening_Balance+Full_Month_Plan-Actual_Usage)/Avg_Consumption), 0)"
                        ],
                        "Format": [
                            "Number (no decimals)",
                            "Percentage",
                            "Percentage", 
                            "Number (no decimals)",
                            "Number (no decimals)",
                            "Number (no decimals)"
                        ]
                    }
                    
                    formula_df = pd.DataFrame(formula_data)
                    st.dataframe(formula_df, use_container_width=True)
                    
                    st.info("💡 **Tip:** Click on any calculated cell in Excel to view and copy the formula!")
            
        except Exception as e:
            st.error(f"❌ An error occurred while processing your data: {str(e)}")
            st.info("💡 Please check your file formats and try again.")
    
    else:
        # Welcome screen when no files are uploaded
        st.markdown("""
        <div class="upload-card">
            <h2>🎯 Welcome to Advanced Bag Report Analytics</h2>
            <p>Get started by uploading your files using the sidebar controls.</p>
            
            <h3>✨ Features:</h3>
            <ul>
                <li>📊 Interactive charts and visualizations</li>
                <li>🔍 Advanced filtering and sorting options</li>
                <li>📈 Real-time performance metrics</li>
                <li>🧮 Excel formulas for transparency</li>
                <li>💾 Professional report generation</li>
                <li>🎨 Beautiful and responsive design</li>
            </ul>
            
            <h3>📋 Required Files:</h3>
            <ol>
                <li><strong>Input Excel File:</strong> Your main consumption data</li>
                <li><strong>Merge Excel File:</strong> Additional planning information</li>
            </ol>
        </div>
        """, unsafe_allow_html=True)
        
        # Sample data preview
        with st.expander("📖 Sample Data Format", expanded=False):
            st.markdown("""
            ### Expected Input File Structure:
            | Plant Name | Brand Name | Bag Name | Opening Balance | ... |
            |------------|------------|----------|-----------------|-----|
            | Plant ABC  | Brand X    | Bag 1    | 1000           | ... |
            
            ### Expected Merge File Structure:
            | Plant Name | Brand Name | Bag Name | Full Month Plan |
            |------------|------------|----------|-----------------|
            | Plant ABC  | Brand X    | Bag 1    | 500            |
            """)

if __name__ == "__main__":
    main()
