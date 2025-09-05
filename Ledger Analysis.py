import io, re
from datetime import datetime
import numpy as np
import pandas as pd
import pdfplumber
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
NUM_RE = re.compile(r'(?<![A-Za-z])[+-]?\d{1,3}(?:,\d{3})*(?:\.\d+)?')
DATE_AT_START = re.compile(r'^\d{2}\.\d{2}\.\d{4}')
CREDIT_INCLUDE = re.compile(r'(?:/DG/|RQDBN|CREDIT\s*NOTE)', re.IGNORECASE)
CREDIT_EXCLUDE = re.compile(r'(?:PIF|COLL|BANK|NEFT|RTGS|UPI|CASH|/DZ/|ADJUST)', re.IGNORECASE)
SALES_RE = re.compile(r'(?P<date>\d{2}\.\d{2}\.\d{4}).*?Sales of-(?P<ctype>[A-Za-z0-9 /+\-&]+?)\s+'r'(?P<qty>\d+(?:\.\d+)?)\s+'r'(?P<rate>\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s+'r'(?P<debit>\d{1,3}(?:,\d{3})*(?:\.\d{2}))')
def _to_float(s: str):
    if not s: return None
    try: return float(s.replace(",", "").strip())
    except: return None
def _dt(s: str):
    try: return datetime.strptime(s, "%d.%m.%Y")
    except: return None
def _norm_product(raw: str) -> str:
    t = (raw or "").upper()
    if "PPC" in t: return "PPC"
    return "OTHER"
def extract_lines(file_bytes: bytes):
    lines=[]
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for p in pdf.pages:
            txt=p.extract_text(x_tolerance=1) or ""
            for ln in txt.splitlines():
                lines.append(" ".join(ln.split()))
    return lines
def parse_sales(lines):
    rows=[]
    for ln in lines:
        m=SALES_RE.search(ln)
        if not m: continue
        dt=_dt(m.group("date"))
        if not dt: continue
        rows.append(dict(date=dt,product=_norm_product(m.group("ctype")),qty_mt=_to_float(m.group("qty")),debit=_to_float(m.group("debit"))))
    df=pd.DataFrame(rows)
    if not df.empty: df=df.sort_values("date").reset_index(drop=True)
    return df
def parse_credit_notes(lines):
    rows=[]
    for ln in lines:
        if not DATE_AT_START.match(ln): continue
        if CREDIT_EXCLUDE.search(ln): continue
        if not CREDIT_INCLUDE.search(ln): continue
        nums=NUM_RE.findall(ln)
        if len(nums)<2: continue
        floats=[_to_float(x) for x in nums if _to_float(x) is not None]
        if len(floats)<2: continue
        amt=floats[-2]   # credit column
        dt=_dt(ln[:10])
        if dt and amt is not None:
            rows.append(dict(date=dt, credit=amt))
    return pd.DataFrame(rows)
def weighted_price(debit, qty_mt): 
    return debit/(qty_mt*20.0) if qty_mt else np.nan
def monthly_ppc(df_sales, df_credit):
    ppc=df_sales[df_sales["product"]=="PPC"].copy()
    if ppc.empty: return pd.DataFrame()
    ppc["ym"]=ppc["date"].dt.to_period("M").astype(str)
    agg=ppc.groupby("ym",as_index=False).agg(qty=("qty_mt","sum"), debit=("debit","sum"))
    agg["Price/Bag"]=agg.apply(lambda r: weighted_price(r["debit"],r["qty"]), axis=1)
    if not df_credit.empty:
        df_credit=df_credit.copy()
        df_credit["ym"]=df_credit["date"].dt.to_period("M").astype(str)
        cn=df_credit.groupby("ym",as_index=False).agg(Discount=("credit","sum"))
    else:
        cn=pd.DataFrame({"ym":[],"Discount":[]})
    out=agg.merge(cn,on="ym",how="left").fillna({"Discount":0.0})
    out=out.rename(columns={"ym":"Year/Month","qty":"QTY(MT) [PPC]"})
    total_qty=out["QTY(MT) [PPC]"].sum()
    total_debit=agg["debit"].sum()
    price=weighted_price(total_debit,total_qty)
    grand=pd.DataFrame([{"Year/Month":"Grand Total","QTY(MT) [PPC]":total_qty,"Price/Bag":price,"Discount":out["Discount"].sum()}])
    return pd.concat([out,grand],ignore_index=True)
def export_excel(report_df, all_qty, district, period, company="JKS"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right")
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    blue = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),top=Side(style='thin'), bottom=Side(style='thin'))
    ws.merge_cells("A1:D1")
    ws["A1"] = f"{company}\n{district}({period})"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A1"].alignment = center
    headers = ["Year/Month", "QTY(MT) [PPC]", "Price/Bag", "Discount"]
    ws.append(headers)
    for c in "ABCD":
        ws[f"{c}2"].font = bold
        ws[f"{c}2"].alignment = center
        ws[f"{c}2"].border = border
    for r in dataframe_to_rows(report_df, index=False, header=False):
        ws.append(r)
    for row in ws.iter_rows(min_row=3, max_row=2+len(report_df), min_col=1, max_col=4):
        for cell in row:
            cell.border = border
            cell.alignment = right if cell.column>1 else center
    for r in range(3, 3+len(report_df)):
        if "Grand" in str(ws[f"A{r}"].value):
            for c in "ABCD":
                ws[f"{c}{r}"].fill = yellow
                ws[f"{c}{r}"].font = bold
    start = 4+len(report_df)
    ws[f"A{start}"] = "Total Qty(All products combined)"
    ws[f"B{start}"] = all_qty  # <- FIXED: uses total qty of all cement types
    ws[f"C{start}"] = "Discount/Bag"
    ws[f"D{start}"] = round(report_df["Discount"].sum()/(all_qty*20),2) if all_qty else 0
    ws[f"A{start}"].font = bold; ws[f"C{start}"].font = bold
    ws[f"C{start+2}"] = "NOD"
    ws[f"D{start+2}"] = round(report_df["Price/Bag"].iloc[-1],2)
    ws[f"C{start+2}"].fill = blue; ws[f"D{start+2}"].fill = blue
    ws[f"C{start+3}"] = f"{company}-Wonder"
    ws[f"D{start+3}"] = "Diff"
    ws[f"C{start+3}"].fill = blue; ws[f"D{start+3}"].fill = blue
    for col in ["A","B","C","D"]:
        ws.column_dimensions[col].width = 22
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
st.title("📊 PPC Ledger Monthly Report with Excel Export")
files=st.file_uploader("Upload ledger PDFs", type="pdf", accept_multiple_files=True)
district=st.text_input("District Name", "Jodhpur")
if files:
    all_sales, all_cn=[],[]
    for f in files:
        lines=extract_lines(f.read())
        all_sales.append(parse_sales(lines))
        all_cn.append(parse_credit_notes(lines))
    df_sales=pd.concat(all_sales,ignore_index=True) if all_sales else pd.DataFrame()
    df_credit=pd.concat(all_cn,ignore_index=True) if all_cn else pd.DataFrame()
    if df_sales.empty: 
        st.error("No PPC sales found."); st.stop()
    report=monthly_ppc(df_sales, df_credit)
    all_qty=df_sales["qty_mt"].sum()
    min_date,max_date=df_sales["date"].min(),df_sales["date"].max()
    period=f"{min_date.strftime('%b\'%y')}–{max_date.strftime('%b\'%y')}"
    st.subheader(f"Report for {district} ({period})")
    st.dataframe(report.style.format({"QTY(MT) [PPC]":"{:,.2f}","Price/Bag":"₹{:,.2f}","Discount":"₹{:,.0f}"}),use_container_width=True)
    excel_bytes=export_excel(report, all_qty, district, period, company="JKS")
    st.download_button("📥 Download Excel Report",data=excel_bytes,file_name=f"report_{district}.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
