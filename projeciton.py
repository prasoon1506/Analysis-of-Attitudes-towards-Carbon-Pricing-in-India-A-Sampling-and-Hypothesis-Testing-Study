import streamlit as st
import pandas as pd
import numpy as np
from sklearn.ensemble import RandomForestRegressor
from sklearn.preprocessing import LabelEncoder
from sklearn.metrics import mean_absolute_percentage_error
import warnings
from openpyxl import load_workbook
import plotly.express as px
import plotly.graph_objects as go
warnings.filterwarnings('ignore')
def read_excel_skip_hidden(uploaded_file):
    wb = load_workbook(uploaded_file)
    ws = wb.active
    hidden_rows = [i + 1 for i in range(ws.max_row) if ws.row_dimensions[i + 1].hidden]
    df = pd.read_excel(uploaded_file,skiprows=hidden_rows)
    return df
def calculate_trend_prediction(features, growth_weights):
    weighted_growth = sum(features[month] * weight for month, weight in growth_weights.items()) / sum(growth_weights.values())
    return features['sales_Oct'] * weighted_growth
def prepare_features_for_optimization(df, target_month='Oct'):
    features = pd.DataFrame()
    if target_month == 'Oct':
        training_months = ['Apr', 'May', 'June', 'July', 'Aug', 'Sep']
        prev_year_months = ['Sep', 'Oct']
        growth_months = training_months[1:]
    else:  # For November predictions
        training_months = ['Apr', 'May', 'June', 'July', 'Aug', 'Sep', 'Oct']
        prev_year_months = ['Sep', 'Oct', 'Nov']
        growth_months = training_months[1:]
    for month in training_months:
        features[f'sales_{month}'] = df[f'Monthly Achievement({month})']
    for month in prev_year_months:
        features[f'prev_{month.lower()}'] = df[f'Total {month} 2023']
    for month in training_months:
        features[f'month_target_{month}'] = df[f'Month Tgt ({month})']
        features[f'ags_target_{month}'] = df[f'AGS Tgt ({month})']
    features['avg_monthly_sales'] = features[[f'sales_{m}' for m in training_months]].mean(axis=1)
    for i in range(len(growth_months)):
        curr_month = growth_months[i]
        prev_month = training_months[i]
        features[f'growth_{curr_month}'] = (features[f'sales_{curr_month}'] / features[f'sales_{prev_month}'])
    if target_month == 'Oct':
        features['yoy_sep_growth'] = features['sales_Sep'] / features['prev_sep']
        features['target_achievement_rate'] = features['sales_Sep'] / features['month_target_Sep']
    else:
        features['yoy_sep_growth'] = features['sales_Sep'] / features['prev_sep']
        features['yoy_oct_growth'] = features['sales_Oct'] / features['prev_oct']
        features['yoy_weighted_growth'] = (features['yoy_sep_growth'] * 0.4 + features['yoy_oct_growth'] * 0.6)
        features['target_achievement_rate'] = features['sales_Oct'] / features['month_target_Oct']
    return features
def generate_predictions_for_optimization(features, df, growth_weights, target_month='Oct'):
    feature_cols = [col for col in features.columns if col not in ['avg_monthly_sales', 'yoy_sep_growth', 'yoy_oct_growth', 'yoy_weighted_growth','target_achievement_rate'] and not col.startswith('growth_')]
    rf_model = RandomForestRegressor(n_estimators=100, random_state=42)
    if target_month == 'Oct':
        rf_model.fit(features[feature_cols], features['sales_Sep'])
        rf_prediction = rf_model.predict(features[feature_cols])
    else:
        rf_model.fit(features[feature_cols], features['sales_Oct'])
        rf_prediction = rf_model.predict(features[feature_cols])
    if target_month == 'Oct':
        yoy_prediction = df['Total Oct 2023'] * features['yoy_sep_growth']
    else:
        yoy_prediction = df['Total Nov 2023'] * features['yoy_weighted_growth']
    if target_month == 'Oct':
        last_month_sales = features['sales_Sep']
    else:
        last_month_sales = features['sales_Oct']
    trend_prediction = calculate_trend_prediction(features, growth_weights) * last_month_sales
    target_based_prediction = features['avg_monthly_sales'] * features['target_achievement_rate']
    return rf_prediction, yoy_prediction, trend_prediction, target_based_prediction
def objective_function(weights, *args):
    rf_pred, yoy_pred, trend_pred, target_pred, actual = args
    final_prediction = (weights[0] * rf_pred +weights[1] * yoy_pred +weights[2] * trend_pred +weights[3] * target_pred)
    return mean_squared_error(actual, final_prediction, squared=False)
def find_optimal_weights(df, zone, brand, growth_weights):
    df_filtered = df[(df['Zone'] == zone) & (df['Brand'] == brand)]
    if len(df_filtered) == 0:
        return None
    features = prepare_features_for_optimization(df_filtered, target_month='Oct')
    rf_pred, yoy_pred, trend_pred, target_pred = generate_predictions_for_optimization(features, df_filtered, growth_weights, target_month='Oct')
    actual_oct = df_filtered['Monthly Achievement(Oct)']
    initial_weights = np.array([0.25, 0.25, 0.25, 0.25])
    bounds = [(0, 1)] * 4
    constraint = {'type': 'eq', 'fun': lambda x: np.sum(x) - 1}
    result = minimize(objective_function,initial_weights,args=(rf_pred, yoy_pred, trend_pred, target_pred, actual_oct),bounds=bounds,constraints=constraint,method='SLSQP')
    return {'rf': result.x[0],'yoy': result.x[1],'trend': result.x[2],'target': result.x[3]}
@st.cache_data
def generate_all_optimal_weights(df, growth_weights):
    optimal_weights = {}
    with st.spinner("Generating optimal weights based on October predictions..."):
        total = len(df['Zone'].unique()) * len(df['Brand'].unique())
        progress_bar = st.progress(0)
        counter = 0
        for zone in df['Zone'].unique():
            optimal_weights[zone] = {}
            for brand in df[df['Zone'] == zone]['Brand'].unique():
                weights = find_optimal_weights(df, zone, brand, growth_weights)
                if weights:
                    optimal_weights[zone][brand] = weights
                counter += 1
                progress_bar.progress(counter / total)
    return optimal_weights
def main():
    st.set_page_config(page_title="Sales Forecasting Model", layout="wide")
    st.markdown("""<style>.stApp {max-width: 1200px;margin: 0 auto;}.stButton button {width: 100%;}.stAlert {padding: 1rem;margin: 1rem 0;}</style>""", unsafe_allow_html=True)
    st.title("📈 Sales Forecasting Model")
    uploaded_file = st.file_uploader("Upload Excel File", type=['xlsx'])
    if uploaded_file is not None:
        try:
            df = read_excel_skip_hidden(uploaded_file)
            default_growth_weights = {'growth_May': 0.05, 'growth_June': 0.1, 'growth_July': 0.15,'growth_Aug': 0.2, 'growth_Sep': 0.25}
            optimal_weights = generate_all_optimal_weights(df, default_growth_weights)
            col1, col2 = st.columns(2)
            with col1:
                st.subheader("Zone Selection")
                zones = sorted(df['Zone'].unique())
                selected_zone = st.selectbox('Select Zone', zones)
            with col2:
                st.subheader("Brand Selection")
                brands = sorted(df[df['Zone'] == selected_zone]['Brand'].unique())
                selected_brand = st.selectbox('Select Brand', brands)
            tab1, tab2 = st.tabs(["Growth Weights", "Method Weights"])
            with tab1:
                st.subheader("Configure Growth Weights")
                growth_weights = {}
                for month, default_weight in default_growth_weights.items():
                    growth_weights[month] = st.slider(f"{month.replace('growth_', '')} Growth Weight",0.0, 1.0, default_weight, 0.05)
                if abs(sum(growth_weights.values()) - 1.0) > 0.01:
                    st.warning("⚠️ Growth weights should sum to 1")
            with tab2:
                st.subheader("Configure Method Weights")
                use_optimal = st.checkbox("Use recommended weights (based on October predictions)")
                method_weights = {}
                if use_optimal and selected_zone in optimal_weights and selected_brand in optimal_weights[selected_zone]:
                    recommended_weights = optimal_weights[selected_zone][selected_brand]
                    for method, weight in recommended_weights.items():
                        method_weights[method] = st.slider(f"{method.upper()} Weight",0.0, 1.0, float(weight), 0.05)
                else:
                    default_method_weights = {'rf': 0.4, 'yoy': 0.1, 'trend': 0.4, 'target': 0.1}
                    for method, default_weight in default_method_weights.items():
                        method_weights[method] = st.slider(f"{method.upper()} Weight",0.0, 1.0, default_weight, 0.05)
                if abs(sum(method_weights.values()) - 1.0) > 0.01:
                    st.warning("⚠️ Method weights should sum to 1")
            if st.button("Generate Predictions", type="primary"):
                if abs(sum(growth_weights.values()) - 1.0) > 0.01 or abs(sum(method_weights.values()) - 1.0) > 0.01:
                    st.error("Please adjust weights to sum to 1 before generating predictions")
                else:
                    with st.spinner("Generating predictions..."):
                        predictions = predict_november_sales(df, selected_zone, selected_brand,growth_weights, method_weights)
                        if predictions is not None:
                            chart_col1, chart_col2 = st.columns(2)
                            fig_methods, fig_gauge = create_prediction_charts(predictions)
                            with chart_col1:
                                st.plotly_chart(fig_methods, use_container_width=True)
                            with chart_col2:
                                st.plotly_chart(fig_gauge, use_container_width=True)
                            st.subheader("Summary Metrics")
                            metric_col1, metric_col2, metric_col3 = st.columns(3)
                            with metric_col1:
                                st.metric("Average Prediction", f"₹{predictions['Final_Prediction'].mean():,.2f}")
                            with metric_col2:
                                st.metric("Minimum Prediction", f"₹{predictions['Final_Prediction'].min():,.2f}")
                            with metric_col3:
                                st.metric("Maximum Prediction", f"₹{predictions['Final_Prediction'].max():,.2f}")
                            st.subheader("Detailed Predictions")
                            styled_predictions = predictions.copy()
                            numeric_cols = predictions.select_dtypes(include=['float64', 'int64']).columns
                            for col in numeric_cols:
                                styled_predictions[col] = styled_predictions[col].map('{:,.2f}'.format)
                            st.dataframe(styled_predictions)
                            csv = predictions.to_csv(index=False)
                            st.download_button(label="Download Predictions as CSV",data=csv,file_name="sales_predictions.csv",mime="text/csv")
        except Exception as e:
            st.error(f"An error occurred: {str(e)}")
            st.info("Please make sure you've uploaded a valid Excel file with the correct format")
if __name__ == "__main__":
    main()
