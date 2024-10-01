import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
from collections import OrderedDict
import plotly.express as px
import plotly.graph_objects as go
from statsmodels.tsa.arima.model import ARIMA
from scipy import stats
import base64
from io import BytesIO
import statsmodels.api as sm
from statsmodels.stats.diagnostic import het_breuschpagan, acorr_ljungbox
from statsmodels.stats.stattools import durbin_watson
from statsmodels.stats.outliers_influence import variance_inflation_factor
from statsmodels.tsa.stattools import adfuller
from statsmodels.graphics.tsaplots import plot_acf, plot_pacf
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler, PolynomialFeatures
from sklearn.linear_model import LinearRegression, Ridge, Lasso
from sklearn.tree import DecisionTreeRegressor
from sklearn.ensemble import RandomForestRegressor
from sklearn.svm import SVR
from sklearn.cluster import KMeans
from sklearn.decomposition import PCA
from sklearn.metrics import mean_squared_error, r2_score
import seaborn as sns
import matplotlib.pyplot as plt
from scipy.stats import jarque_bera, kurtosis, skew
from statsmodels.stats.stattools import omni_normtest

def excel_editor_and_analyzer():
    st.set_page_config(layout="wide")
    st.title("Advanced Excel Editor and Analyzer")
    
    apply_custom_css()
    
    tab1, tab2 = st.tabs(["Excel Editor", "Data Analyzer"])
    
    with tab1:
        excel_editor()
    
    with tab2:
        data_analyzer()

def apply_custom_css():
    st.markdown("""
    <style>
        .stApp {
            background-color: #f0f2f6;
        }
        .excel-table {
            border-collapse: collapse;
            width: 100%;
            font-family: Arial, sans-serif;
        }
        .excel-table th, .excel-table td {
            border: 1px solid #b0b0b0;
            padding: 8px;
            text-align: left;
        }
        .excel-table tr:nth-child(even) {
            background-color: #f8f8f8;
        }
        .excel-table th {
            padding-top: 12px;
            padding-bottom: 12px;
            background-color: #4CAF50;
            color: white;
        }
        .stButton>button {
            background-color: #4CAF50;
            color: white;
            border: none;
            padding: 10px 24px;
            text-align: center;
            text-decoration: none;
            display: inline-block;
            font-size: 16px;
            margin: 4px 2px;
            cursor: pointer;
            border-radius: 4px;
        }
        .stTextInput>div>div>input {
            color: #4CAF50;
        }
        .stSelectbox>div>div>select {
            color: #4CAF50;
        }
        .stMultiSelect>div>div>select {
            color: #4CAF50;
        }
    </style>
    """, unsafe_allow_html=True)

def excel_editor():
    st.header("Excel Editor")
    def create_excel_structure_html(sheet, max_rows=5):
        html = "<table class='excel-table'>"
        merged_cells = sheet.merged_cells.ranges

        for idx, row in enumerate(sheet.iter_rows(max_row=max_rows)):
            html += "<tr>"
            for cell in row:
                merged = False
                for merged_range in merged_cells:
                    if cell.coordinate in merged_range:
                        if cell.coordinate == merged_range.start_cell.coordinate:
                            rowspan = min(merged_range.max_row - merged_range.min_row + 1, max_rows - idx)
                            colspan = merged_range.max_col - merged_range.min_col + 1
                            html += f"<td rowspan='{rowspan}' colspan='{colspan}'>{cell.value}</td>"
                        merged = True
                        break
                if not merged:
                    html += f"<td>{cell.value}</td>"
            html += "</tr>"
        html += "</table>"
        return html

    # Function to get merged column groups
    def get_merged_column_groups(sheet):
        merged_groups = {}
        for merged_range in sheet.merged_cells.ranges:
            if merged_range.min_row == 1:  # Only consider merged cells in the first row (header)
                main_col = sheet.cell(1, merged_range.min_col).value
                merged_groups[main_col] = list(range(merged_range.min_col, merged_range.max_col + 1))
        return merged_groups

    # File uploader
    uploaded_file = st.file_uploader("Choose an Excel file", type="xlsx")

    if uploaded_file is not None:
        # Read Excel file
        excel_file = openpyxl.load_workbook(uploaded_file)
        sheet = excel_file.active

        # Display original Excel structure (first 5 rows)
        st.subheader("Original Excel Structure (First 5 Rows)")
        excel_html = create_excel_structure_html(sheet, max_rows=5)
        st.markdown(excel_html, unsafe_allow_html=True)

        # Get merged column groups
        merged_groups = get_merged_column_groups(sheet)

        # Create a list of column headers, considering merged cells
        column_headers = []
        column_indices = OrderedDict()  # To store the column indices for each header
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, col).value
            if cell_value is not None:
                column_headers.append(cell_value)
                if cell_value not in column_indices:
                    column_indices[cell_value] = []
                column_indices[cell_value].append(col - 1)  # pandas uses 0-based index
            else:
                # If the cell is empty, it's part of a merged cell, so use the previous header
                prev_header = column_headers[-1]
                column_headers.append(prev_header)
                column_indices[prev_header].append(col - 1)

        # Read as pandas DataFrame using the correct column headers
        df = pd.read_excel(uploaded_file, header=None, names=column_headers)
        df = df.iloc[1:]  # Remove the first row as it's now our header

        # Column selection for deletion
        st.subheader("Select columns to delete")
        all_columns = list(column_indices.keys())  # Use OrderedDict keys to maintain order
        cols_to_delete = st.multiselect("Choose columns to remove", all_columns)
        
        if cols_to_delete:
            columns_to_remove = []
            for col in cols_to_delete:
                columns_to_remove.extend(column_indices[col])
            
            df = df.drop(df.columns[columns_to_remove], axis=1)
            st.success(f"Deleted columns: {', '.join(cols_to_delete)}")

        # Row deletion
        st.subheader("Delete rows")
        num_rows = st.number_input("Enter the number of rows to delete from the start", min_value=0, max_value=len(df)-1, value=0)
        
        if num_rows > 0:
            df = df.iloc[num_rows:]
            st.success(f"Deleted first {num_rows} rows")
        
        # Display editable dataframe
        st.subheader("Edit Data")
        st.write("You can edit individual cell values directly in the table below:")
        
        # Replace NaN values with None and convert dataframe to a dictionary
        df_dict = df.where(pd.notnull(df), None).to_dict('records')
        
        # Use st.data_editor with the processed dictionary
        edited_data = st.data_editor(df_dict)
        
        # Convert edited data back to dataframe
        edited_df = pd.DataFrame(edited_data)
        st.subheader("Edited Data")
        st.dataframe(edited_df)
        
        # Download button
        def get_excel_download_link(df):
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)
            excel_data = output.getvalue()
            b64 = base64.b64encode(excel_data).decode()
            return f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="edited_file.xlsx">Download Edited Excel File</a>'
        
        st.markdown(get_excel_download_link(edited_df), unsafe_allow_html=True)

        # New button to upload edited file to Home
        if st.button("Upload Edited File to Home"):
            # Save the edited DataFrame to session state
            st.session_state.edited_df = edited_df
            st.session_state.edited_file_name = "edited_" + uploaded_file.name
            st.success("Edited file has been uploaded to Home. Please switch to the Home tab to see the uploaded file.")

    else:
        st.info("Please upload an Excel file to begin editing.")

def data_analyzer():
    st.header("Advanced Data Analyzer")
    
    uploaded_file = st.file_uploader("Choose an Excel file for analysis", type="xlsx", key="analyser")
    
    if uploaded_file is not None:
        df = pd.read_excel(uploaded_file)
        
        st.write("Dataset Information:")
        st.write(f"Number of rows: {df.shape[0]}")
        st.write(f"Number of columns: {df.shape[1]}")
        
        numeric_columns = df.select_dtypes(include=['float64', 'int64']).columns
        categorical_columns = df.select_dtypes(include=['object']).columns
        
        analysis_type = st.selectbox("Select analysis type", ["Univariate Analysis", "Bivariate Analysis", "Regression Analysis", "Machine Learning Models", "Advanced Statistics"])
        
        if analysis_type == "Univariate Analysis":
            univariate_analysis(df, numeric_columns, categorical_columns)
        elif analysis_type == "Bivariate Analysis":
            bivariate_analysis(df, numeric_columns)
        elif analysis_type == "Regression Analysis":
            regression_analysis(df, numeric_columns, categorical_columns)
        elif analysis_type == "Machine Learning Models":
            machine_learning_models(df, numeric_columns, categorical_columns)
        elif analysis_type == "Advanced Statistics":
            advanced_statistics(df, numeric_columns)

def univariate_analysis(df, numeric_columns, categorical_columns):
    st.subheader("Univariate Analysis")
    
    column = st.selectbox("Select a column for analysis", numeric_columns.tolist() + categorical_columns.tolist())
    
    if column in numeric_columns:
        st.write(df[column].describe())
        
        col1, col2 = st.columns(2)
        
        with col1:
            fig = go.Figure()
            fig.add_trace(go.Histogram(x=df[column], name="Histogram"))
            fig.update_layout(title=f"Histogram for {column}")
            st.plotly_chart(fig, use_container_width=True)
        
        with col2:
            fig = go.Figure()
            fig.add_trace(go.Box(y=df[column], name="Box Plot"))
            fig.update_layout(title=f"Box Plot for {column}")
            st.plotly_chart(fig, use_container_width=True)
        
        col3, col4 = st.columns(2)
        
        with col3:
            fig = go.Figure()
            fig.add_trace(go.Violin(y=df[column], box_visible=True, line_color='black', meanline_visible=True, fillcolor='lightseagreen', opacity=0.6, x0=column))
            fig.update_layout(title=f"Violin Plot for {column}")
            st.plotly_chart(fig, use_container_width=True)
        
        with col4:
            fig = px.line(df, y=column, title=f"Line Plot for {column}")
            st.plotly_chart(fig, use_container_width=True)
        
        # Additional statistics
        st.subheader("Additional Statistics")
        col5, col6, col7 = st.columns(3)
        with col5:
            st.metric("Skewness", f"{skew(df[column]):.4f}")
        with col6:
            st.metric("Kurtosis", f"{kurtosis(df[column]):.4f}")
        with col7:
            st.metric("Coefficient of Variation", f"{df[column].std() / df[column].mean():.4f}")
        
    else:
        st.write(df[column].value_counts())
        col1, col2 = st.columns(2)
        
        with col1:
            fig = px.bar(df[column].value_counts(), title=f"Bar Plot for {column}")
            st.plotly_chart(fig, use_container_width=True)
        
        with col2:
            fig = px.pie(df, names=column, title=f"Pie Chart for {column}")
            st.plotly_chart(fig, use_container_width=True)

def bivariate_analysis(df, numeric_columns):
    st.subheader("Bivariate Analysis")
    
    x_col = st.selectbox("Select X-axis variable", numeric_columns)
    y_col = st.selectbox("Select Y-axis variable", numeric_columns)
    
    chart_type = st.selectbox("Select chart type", ["Scatter", "Line", "Bar", "Box", "Violin", "3D Scatter", "Heatmap"])
    
    if chart_type == "Scatter":
        fig = px.scatter(df, x=x_col, y=y_col, title=f"{y_col} vs {x_col}")
    elif chart_type == "Line":
        fig = px.line(df, x=x_col, y=y_col, title=f"{y_col} vs {x_col}")
    elif chart_type == "Bar":
        fig = px.bar(df, x=x_col, y=y_col, title=f"{y_col} vs {x_col}")
    elif chart_type == "Box":
        fig = px.box(df, x=x_col, y=y_col, title=f"{y_col} vs {x_col}")
    elif chart_type == "Violin":
        fig = px.violin(df, x=x_col, y=y_col, title=f"{y_col} vs {x_col}")
    elif chart_type == "3D Scatter":
        z_col = st.selectbox("Select Z-axis variable", numeric_columns)
        fig = px.scatter_3d(df, x=x_col, y=y_col, z=z_col, title=f"3D Scatter Plot")
    elif chart_type == "Heatmap":
        corr_matrix = df[numeric_columns].corr()
        fig = px.imshow(corr_matrix, title="Correlation Heatmap")
    
    st.plotly_chart(fig, use_container_width=True)
    
    correlation = df[[x_col, y_col]].corr().iloc[0, 1]
    st.write(f"Correlation between {x_col} and {y_col}: {correlation:.4f}")
    
    # Add correlation interpretation
    st.subheader("Correlation Interpretation")
    st.write("""
    The correlation coefficient ranges from -1 to 1:
    - 1: Perfect positive correlation
    - 0: No correlation
    - -1: Perfect negative correlation
    
    Interpretation:
    - 0.00 to 0.19: Very weak correlation
    - 0.20 to 0.39: Weak correlation
    - 0.40 to 0.59: Moderate correlation
    - 0.60 to 0.79: Strong correlation
    - 0.80 to 1.00: Very strong correlation
    """)
    
    # Add correlation formula
    st.latex(r'''
    r = \frac{\sum_{i=1}^{n} (x_i - \bar{x})(y_i - \bar{y})}{\sqrt{\sum_{i=1}^{n} (x_i - \bar{x})^2} \sqrt{\sum_{i=1}^{n} (y_i - \bar{y})^2}}
    ''')
    st.write("Where:")
    st.write("- r is the correlation coefficient")
    st.write("- x_i and y_i are individual sample points")
    st.write("- x̄ and ȳ are the sample means")

def regression_analysis(df, numeric_columns, categorical_columns):
    st.subheader("Regression Analysis")
    
    regression_type = st.selectbox("Select regression type", ["Simple Linear", "Multiple Linear", "Polynomial", "Ridge", "Lasso"])
    
    y_col = st.selectbox("Select dependent variable", numeric_columns)
    x_cols = st.multiselect("Select independent variables", numeric_columns.tolist() + categorical_columns.tolist())
    
    if len(x_cols) == 0:
        st.warning("Please select at least one independent variable.")
        return
    
    X = df[x_cols]
    y = df[y_col]
    
    # Handle categorical variables
    X = pd.get_dummies(X, drop_first=True)
    
    if regression_type == "Polynomial":
        degree = st.slider("Select polynomial degree", 1, 5, 2)
        poly = PolynomialFeatures(degree=degree)
        X = poly.fit_transform(X)
    
    X = sm.add_constant(X)
    
    try:
        if regression_type == "Ridge":
            alpha = st.slider("Select alpha for Ridge regression", 0.0, 10.0, 1.0)
            model = sm.OLS(y, X).fit_regularized(alpha=alpha, L1_wt=0)
        elif regression_type == "Lasso":
            alpha = st.slider("Select alpha for Lasso regression", 0.0, 10.0, 1.0)
            model = sm.OLS(y, X).fit_regularized(alpha=alpha, L1_wt=1)
        else:
            model = sm.OLS(y, X).fit()
        
        st.write(model.summary())
        
        # Plot actual vs predicted values
        fig = px.scatter(x=y, y=model.predict(X), labels={'x': 'Actual', 'y': 'Predicted'}, title="Actual vs Predicted Values")
        fig.add_trace(go.Scatter(x=[y.min(), y.max()], y=[y.min(), y.max()], mode='lines', name='y=x'))
        st.plotly_chart(fig, use_container_width=True)
        
        # Residual plot
        residuals = model.resid
        fig = px.scatter(x=model.predict(X), y=residuals, labels={'x': 'Predicted', 'y': 'Residuals'}, title="Residual Plot")
        fig.add_hline(y=0, line_dash="dash", line_color="red")
        st.plotly_chart(fig, use_container_width=True)
        
        # Statistical tests
        st.subheader("Statistical Tests")
        
        # Normality test (Jarque-Bera)
        jb_statistic, jb_p_value = jarque_bera(residuals)
        st.write(f"Jarque-Bera Test for Normality: statistic = {jb_statistic:.4f}, p-value = {jb_p_value:.4f}")
        st.write(f"{'Reject' if jb_p_value < 0.05 else 'Fail to reject'} the null hypothesis of normality at 5% significance level.")
        
        # Heteroscedasticity test (Breusch-Pagan)
        _, bp_p_value, _, _ = het_breuschpagan(residuals, model.model.exog)
        st.write(f"Breusch-Pagan Test for Heteroscedasticity: p-value = {bp_p_value:.4f}")
        st.write(f"{'Reject' if bp_p_value < 0.05 else 'Fail to reject'} the null hypothesis of homoscedasticity at 5% significance level.")
        dw_statistic = durbin_watson(residuals)
        st.write(f"Durbin-Watson Test for Autocorrelation: {dw_statistic:.4f}")
        st.write("Values close to 2 suggest no autocorrelation, while values toward 0 or 4 suggest positive or negative autocorrelation.")
        
        # Multicollinearity (VIF)
        vif_data = pd.DataFrame()
        vif_data["Variable"] = X.columns
        vif_data["VIF"] = [variance_inflation_factor(X.values, i) for i in range(X.shape[1])]
        st.write("Variance Inflation Factors (VIF) for Multicollinearity:")
        st.write(vif_data)
        st.write("VIF > 5 suggests high multicollinearity.")
        
        # Add regression formulas and explanations
        st.subheader("Regression Formulas")
        if regression_type == "Simple Linear":
            st.latex(r'y = \beta_0 + \beta_1x + \epsilon')
        elif regression_type == "Multiple Linear":
            st.latex(r'y = \beta_0 + \beta_1x_1 + \beta_2x_2 + ... + \beta_nx_n + \epsilon')
        elif regression_type == "Polynomial":
            st.latex(r'y = \beta_0 + \beta_1x + \beta_2x^2 + ... + \beta_nx^n + \epsilon')
        elif regression_type == "Ridge":
            st.latex(r'\min_{\beta} \sum_{i=1}^n (y_i - \beta_0 - \sum_{j=1}^p \beta_jx_{ij})^2 + \lambda \sum_{j=1}^p \beta_j^2')
        elif regression_type == "Lasso":
            st.latex(r'\min_{\beta} \sum_{i=1}^n (y_i - \beta_0 - \sum_{j=1}^p \beta_jx_{ij})^2 + \lambda \sum_{j=1}^p |\beta_j|')
        
        st.write("Where:")
        st.write("- y is the dependent variable")
        st.write("- x, x_1, x_2, ..., x_n are independent variables")
        st.write("- β_0, β_1, β_2, ..., β_n are regression coefficients")
        st.write("- ε is the error term")
        st.write("- λ is the regularization parameter (for Ridge and Lasso)")
        
    except Exception as e:
        st.error(f"An error occurred during regression analysis: {str(e)}")
        st.write("This error might be due to multicollinearity, insufficient data, or other issues in the dataset.")
        st.write("Try selecting different variables or using a different regression type.")

def machine_learning_models(df, numeric_columns, categorical_columns):
    st.subheader("Machine Learning Models")
    
    model_type = st.selectbox("Select model type", ["Supervised", "Unsupervised"])
    
    if model_type == "Supervised":
        supervised_models(df, numeric_columns, categorical_columns)
    else:
        unsupervised_models(df, numeric_columns)

def supervised_models(df, numeric_columns, categorical_columns):
    st.write("Supervised Learning Models")
    
    y_col = st.selectbox("Select target variable", numeric_columns)
    x_cols = st.multiselect("Select features", numeric_columns.tolist() + categorical_columns.tolist())
    
    if len(x_cols) == 0:
        st.warning("Please select at least one feature.")
        return
    
    X = df[x_cols]
    y = df[y_col]
    
    # Handle categorical variables
    X = pd.get_dummies(X, drop_first=True)
    
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)
    
    scaler = StandardScaler()
    X_train_scaled = scaler.fit_transform(X_train)
    X_test_scaled = scaler.transform(X_test)
    
    models = {
        "Linear Regression": LinearRegression(),
        "Decision Tree": DecisionTreeRegressor(),
        "Random Forest": RandomForestRegressor(),
        "SVR": SVR()
    }
    
    selected_model = st.selectbox("Select a model", list(models.keys()))
    
    try:
        model = models[selected_model]
        model.fit(X_train_scaled, y_train)
        
        y_pred = model.predict(X_test_scaled)
        
        mse = mean_squared_error(y_test, y_pred)
        r2 = r2_score(y_test, y_pred)
        
        st.write(f"Mean Squared Error: {mse:.4f}")
        st.write(f"R-squared Score: {r2:.4f}")
        
        fig = px.scatter(x=y_test, y=y_pred, labels={'x': 'Actual', 'y': 'Predicted'}, title="Actual vs Predicted Values")
        fig.add_trace(go.Scatter(x=[y_test.min(), y_test.max()], y=[y_test.min(), y_test.max()], mode='lines', name='y=x'))
        st.plotly_chart(fig, use_container_width=True)
        
        # Feature importance (for tree-based models)
        if selected_model in ["Decision Tree", "Random Forest"]:
            feature_importance = pd.DataFrame({
                'feature': X.columns,
                'importance': model.feature_importances_
            }).sort_values('importance', ascending=False)
            
            st.write("Feature Importance:")
            fig = px.bar(feature_importance, x='feature', y='importance', title="Feature Importance")
            st.plotly_chart(fig, use_container_width=True)
        
        # Add model formulas and explanations
        st.subheader("Model Formulas and Explanations")
        if selected_model == "Linear Regression":
            st.latex(r'y = \beta_0 + \beta_1x_1 + \beta_2x_2 + ... + \beta_nx_n')
            st.write("Linear Regression finds the best-fitting linear relationship between the target variable and the features.")
        elif selected_model == "Decision Tree":
            st.write("Decision Trees make predictions by learning decision rules inferred from the data features.")
            st.image("https://scikit-learn.org/stable/_images/iris_dtc.png", caption="Example of a Decision Tree")
        elif selected_model == "Random Forest":
            st.write("Random Forest is an ensemble of Decision Trees, where each tree is trained on a random subset of the data and features.")
            st.image("https://scikit-learn.org/stable/_images/plot_forest_importances_faces_001.png", caption="Example of Random Forest Feature Importance")
        elif selected_model == "SVR":
            st.latex(r'\min_{w, b, \xi} \frac{1}{2} \|w\|^2 + C \sum_{i=1}^n \xi_i')
            st.write("Support Vector Regression (SVR) finds a function that deviates from y by a value no greater than ε for each training point x.")
    
    except Exception as e:
        st.error(f"An error occurred during model training: {str(e)}")
        st.write("This error might be due to insufficient data, incompatible data types, or other issues in the dataset.")
        st.write("Try selecting different variables or using a different model.")

def unsupervised_models(df, numeric_columns):
    st.write("Unsupervised Learning Models")
    
    x_cols = st.multiselect("Select features for clustering", numeric_columns)
    
    if len(x_cols) == 0:
        st.warning("Please select at least one feature.")
        return
    
    X = df[x_cols]
    
    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X)
    
    n_clusters = st.slider("Select number of clusters", 2, 10, 3)
    
    try:
        kmeans = KMeans(n_clusters=n_clusters, random_state=42)
        cluster_labels = kmeans.fit_predict(X_scaled)
        
        df_clustered = df.copy()
        df_clustered['Cluster'] = cluster_labels
        
        if len(x_cols) >= 2:
            fig = px.scatter(df_clustered, x=x_cols[0], y=x_cols[1], color='Cluster', title="K-means Clustering")
            st.plotly_chart(fig, use_container_width=True)
        
        st.write("Cluster Centers:")
        cluster_centers = scaler.inverse_transform(kmeans.cluster_centers_)
        st.write(pd.DataFrame(cluster_centers, columns=x_cols))
        
        # Elbow method for optimal number of clusters
        inertias = []
        k_range = range(1, 11)
        for k in k_range:
            kmeans = KMeans(n_clusters=k, random_state=42)
            kmeans.fit(X_scaled)
            inertias.append(kmeans.inertia_)
        
        fig = px.line(x=k_range, y=inertias, title="Elbow Method for Optimal k",
                      labels={'x': 'Number of Clusters (k)', 'y': 'Inertia'})
        st.plotly_chart(fig, use_container_width=True)
        
        # PCA
        st.subheader("Principal Component Analysis (PCA)")
        n_components = st.slider("Select number of components", 2, min(len(x_cols), 10), 2)
        pca = PCA(n_components=n_components)
        pca_result = pca.fit_transform(X_scaled)
        
        df_pca = pd.DataFrame(data=pca_result, columns=[f'PC{i+1}' for i in range(n_components)])
        
        fig = px.scatter(df_pca, x='PC1', y='PC2', title="PCA Visualization")
        st.plotly_chart(fig, use_container_width=True)
        
        explained_variance_ratio = pca.explained_variance_ratio_
        cumulative_variance_ratio = np.cumsum(explained_variance_ratio)
        
        fig = go.Figure()
        fig.add_trace(go.Bar(x=range(1, n_components+1), y=explained_variance_ratio, name='Individual'))
        fig.add_trace(go.Scatter(x=range(1, n_components+1), y=cumulative_variance_ratio, mode='lines+markers', name='Cumulative'))
        fig.update_layout(title='Explained Variance Ratio', xaxis_title='Principal Components', yaxis_title='Explained Variance Ratio')
        st.plotly_chart(fig, use_container_width=True)
        
        st.write("Explained Variance Ratio:")
        st.write(pd.DataFrame({'PC': range(1, n_components+1), 'Explained Variance Ratio': explained_variance_ratio, 'Cumulative Variance Ratio': cumulative_variance_ratio}))
        
        # Add formulas and explanations
        st.subheader("K-means Clustering Formula")
        st.latex(r'\min_{S} \sum_{i=1}^{k} \sum_{x \in S_i} \|x - \mu_i\|^2')
        st.write("Where:")
        st.write("- S is the set of clusters")
        st.write("- k is the number of clusters")
        st.write("- x is a data point")
        st.write("- μ_i is the mean of points in S_i")
        
        st.subheader("PCA Formula")
        st.latex(r'X = U\Sigma V^T')
        st.write("Where:")
        st.write("- X is the original data matrix")
        st.write("- U is the left singular vectors (eigenvectors of XX^T)")
        st.write("- Σ is a diagonal matrix of singular values")
        st.write("- V^T is the right singular vectors (eigenvectors of X^TX)")
    
    except Exception as e:
        st.error(f"An error occurred during unsupervised learning: {str(e)}")
        st.write("This error might be due to insufficient data, incompatible data types, or other issues in the dataset.")
        st.write("Try selecting different variables or adjusting the number of clusters/components.")

def advanced_statistics(df, numeric_columns):
    st.subheader("Advanced Statistics")
    
    column = st.selectbox("Select a column for advanced statistics", numeric_columns)
    
    st.write("Descriptive Statistics:")
    st.write(df[column].describe())
    
    st.subheader("Normality Tests")
    
    # Shapiro-Wilk Test
    shapiro_stat, shapiro_p = stats.shapiro(df[column])
    st.write(f"Shapiro-Wilk Test: statistic = {shapiro_stat:.4f}, p-value = {shapiro_p:.4f}")
    st.write(f"{'Reject' if shapiro_p < 0.05 else 'Fail to reject'} the null hypothesis of normality at 5% significance level.")
    
    # Anderson-Darling Test
    anderson_result = stats.anderson(df[column])
    st.write("Anderson-Darling Test:")
    st.write(f"Statistic: {anderson_result.statistic:.4f}")
    for i in range(len(anderson_result.critical_values)):
        sl, cv = anderson_result.significance_level[i], anderson_result.critical_values[i]
        st.write(f"At {sl}% significance level: critical value = {cv:.4f}")
        if anderson_result.statistic < cv:
            st.write(f"The null hypothesis of normality is not rejected at {sl}% significance level.")
        else:
            st.write(f"The null hypothesis of normality is rejected at {sl}% significance level.")
    
    # Jarque-Bera Test
    jb_stat, jb_p = stats.jarque_bera(df[column])
    st.write(f"Jarque-Bera Test: statistic = {jb_stat:.4f}, p-value = {jb_p:.4f}")
    st.write(f"{'Reject' if jb_p < 0.05 else 'Fail to reject'} the null hypothesis of normality at 5% significance level.")
    
    # Q-Q Plot
    fig, ax = plt.subplots()
    stats.probplot(df[column], dist="norm", plot=ax)
    ax.set_title("Q-Q Plot")
    st.pyplot(fig)
    
    st.subheader("Time Series Analysis")
    
    # Augmented Dickey-Fuller Test for Stationarity
    adf_result = adfuller(df[column])
    st.write("Augmented Dickey-Fuller Test:")
    st.write(f"ADF Statistic: {adf_result[0]:.4f}")
    st.write(f"p-value: {adf_result[1]:.4f}")
    for key, value in adf_result[4].
