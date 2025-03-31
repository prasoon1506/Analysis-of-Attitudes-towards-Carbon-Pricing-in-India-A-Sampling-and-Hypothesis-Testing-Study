from openpyxl import load_workbook 
from openpyxl.styles import PatternFill 
import re 
def is_percentage(value): 
    if isinstance(value, str): 
        cleaned = value.strip().rstrip('%') 
        try: 
            float(cleaned) 
            return True 
        except ValueError: 
            return False 
    return isinstance(value, (int, float)) 
def get_numeric_value(value):  
    if isinstance(value, str):  
        return float(value.strip().rstrip('%')) 
    return float(value) 
def format_excel_file(input_file, output_file): 
    wb = load_workbook(input_file) 
    ws = wb.active 
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid') 
    green_fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid') 
    column_pairs = [ ('G', 'H'), ('I', 'J'), ('K', 'L'), ('M', 'N') ] 
    column_pairs = [ ('M', 'N'), ('Q', 'R'), ('S', 'T'), ('U', 'V') ] 
    for row in range(2, ws.max_row + 1):
        for col1, col2 in column_pairs: 
            cell1 = ws[f'{col1}{row}'] 
            cell2 = ws[f'{col2}{row}'] 
            if cell1.value is None or cell2.value is None: 
                continue 
            if is_percentage(cell1.value) and is_percentage(cell2.value): 
                val1 = get_numeric_value(cell1.value) 
                val2 = get_numeric_value(cell2.value) 
                if val2 < val1: 
                    ws[f'{col2}{row}'].fill = red_fill 
                else: 
                    ws[f'{col2}{row}'].fill = green_fill 
    wb.save(output_file) 
    print(f"File has been processed and saved as {output_file}") 
if __name__ == "__main__": 
    input_file = "/content/Copy of President_Report_YTD.xlsx" 
    output_file = "output.xlsx" 
    format_excel_file(input_file, output_file) 
